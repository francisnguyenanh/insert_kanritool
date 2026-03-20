import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Đường dẫn file
file_old = 'old.xlsx'
file_new = 'new.xlsx'
output_file = 'comparison_result.txt'

# Tạo bản sao để tô màu
file_old_highlighted = 'old_highlighted.xlsx'
file_new_highlighted = 'new_highlighted.xlsx'

# Kiểm tra file tồn tại
if not os.path.exists(file_old):
    print(f"File {file_old} không tồn tại!")
    exit(1)
if not os.path.exists(file_new):
    print(f"File {file_new} không tồn tại!")
    exit(1)

print("Đang so sánh các file Excel và tô màu...")

# Tạo màu vàng
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Copy file gốc để giữ format
shutil.copy2(file_old, file_old_highlighted)
shutil.copy2(file_new, file_new_highlighted)

# Load workbook
wb_old = load_workbook(file_old, data_only=True)
wb_new = load_workbook(file_new, data_only=True)
wb_old_highlighted = load_workbook(file_old_highlighted)
wb_new_highlighted = load_workbook(file_new_highlighted)

sheets_with_differences = []

with open(output_file, 'w', encoding='utf-8') as f:
    sheets_old = wb_old.sheetnames
    sheets_new = wb_new.sheetnames
    if len(sheets_old) != len(sheets_new):
        f.write("Số lượng sheet không giống nhau!\n")
        f.write(f"File old có {len(sheets_old)} sheets: {', '.join(sheets_old)}\n")
        f.write(f"File new có {len(sheets_new)} sheets: {', '.join(sheets_new)}\n\n")
    else:
        f.write("Số lượng sheet giống nhau. Kết quả so sánh:\n\n")
    common_sheets = set(sheets_old) & set(sheets_new)
    only_in_old = set(sheets_old) - set(sheets_new)
    only_in_new = set(sheets_new) - set(sheets_old)
    if only_in_old:
        f.write(f"Sheet chỉ có trong file old: {', '.join(only_in_old)}\n")
    if only_in_new:
        f.write(f"Sheet chỉ có trong file new: {', '.join(only_in_new)}\n")
    if only_in_old or only_in_new:
        f.write("\n")
    for sheet_name in common_sheets:
        print(f"Đang xử lý sheet: {sheet_name}")
        try:
            ws_old = wb_old[sheet_name]
            ws_new = wb_new[sheet_name]
            ws_old_highlighted = wb_old_highlighted[sheet_name]
            ws_new_highlighted = wb_new_highlighted[sheet_name]
            max_row = max(ws_old.max_row, ws_new.max_row)
            max_col = max(ws_old.max_column, ws_new.max_column)
            different_cells = 0
            different_positions = []
            has_differences = False
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    value_old = ws_old.cell(row=row, column=col).value
                    value_new = ws_new.cell(row=row, column=col).value
                    clean_value_old = str(value_old).strip() if value_old is not None else ''
                    clean_value_new = str(value_new).strip() if value_new is not None else ''
                    if clean_value_old != clean_value_new:
                        different_cells += 1
                        has_differences = True
                        try:
                            cell_old = ws_old_highlighted.cell(row=row, column=col)
                            merged_old = False
                            for merged_range in ws_old_highlighted.merged_cells.ranges:
                                if cell_old.coordinate in merged_range:
                                    for merged_cell in merged_range.cells:
                                        ws_old_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                    merged_old = True
                                    break
                            if not merged_old:
                                cell_old.fill = yellow_fill
                            cell_new = ws_new_highlighted.cell(row=row, column=col)
                            merged_new = False
                            for merged_range in ws_new_highlighted.merged_cells.ranges:
                                if cell_new.coordinate in merged_range:
                                    for merged_cell in merged_range.cells:
                                        ws_new_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                    merged_new = True
                                    break
                            if not merged_new:
                                cell_new.fill = yellow_fill
                        except Exception as e:
                            pass
                        if len(different_positions) < 20:
                            col_letter = ws_old.cell(row=row, column=col).column_letter
                            different_positions.append({
                                'position': f"{col_letter}{row}",
                                'value_old': clean_value_old if clean_value_old != '' else "NULL",
                                'value_new': clean_value_new if clean_value_new != '' else "NULL"
                            })
            values_old = set()
            values_new = set()
            non_empty_rows_old = 0
            non_empty_rows_new = 0
            for row in range(1, ws_old.max_row + 1):
                has_data = False
                for col in range(1, ws_old.max_column + 1):
                    value = ws_old.cell(row=row, column=col).value
                    if value is not None and str(value).strip() != '':
                        clean_value = str(value).strip()
                        values_old.add(clean_value)
                        has_data = True
                if has_data:
                    non_empty_rows_old += 1
            for row in range(1, ws_new.max_row + 1):
                has_data = False
                for col in range(1, ws_new.max_column + 1):
                    value = ws_new.cell(row=row, column=col).value
                    if value is not None and str(value).strip() != '':
                        clean_value = str(value).strip()
                        values_new.add(clean_value)
                        has_data = True
                if has_data:
                    non_empty_rows_new += 1
            if has_differences or values_old != values_new:
                sheets_with_differences.append(sheet_name)
                f.write(f"Sheet: {sheet_name}\n")
                f.write(f"Số dòng có dữ liệu trong file old: {non_empty_rows_old}\n")
                f.write(f"Số dòng có dữ liệu trong file new: {non_empty_rows_new}\n")
                f.write(f"Tổng số giá trị khác rỗng trong file old: {len(values_old)}\n")
                f.write(f"Tổng số giá trị khác rỗng trong file new: {len(values_new)}\n")
                f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                if different_positions:
                    f.write(f"Ví dụ {min(len(different_positions), 20)} vị trí khác nhau đầu tiên:\n")
                    for pos in different_positions:
                        f.write(f"  {pos['position']}: old='{pos['value_old']}' vs new='{pos['value_new']}'\n")
                only_in_old = values_old - values_new
                only_in_new = values_new - values_old
                if only_in_old:
                    f.write(f"Giá trị chỉ có trong file old ({len(only_in_old)} giá trị):\n")
                    for value in sorted(list(only_in_old)[:20]):
                        f.write(f"  {value}\n")
                    if len(only_in_old) > 20:
                        f.write(f"  ... và {len(only_in_old) - 20} giá trị khác\n")
                if only_in_new:
                    f.write(f"Giá trị chỉ có trong file new ({len(only_in_new)} giá trị):\n")
                    for value in sorted(list(only_in_new)[:20]):
                        f.write(f"  {value}\n")
                    if len(only_in_new) > 20:
                        f.write(f"  ... và {len(only_in_new) - 20} giá trị khác\n")
                f.write("\n" + "="*50 + "\n")
            else:
                f.write(f"Sheet: {sheet_name}\n")
                f.write("Nội dung giống hệt nhau (bỏ qua dòng trắng và cấu trúc).\n")
                f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                f.write("\n" + "="*50 + "\n")
        except Exception as e:
            f.write(f"Sheet: {sheet_name}\n")
            f.write(f"Lỗi khi xử lý sheet: {e}\n")
            f.write("\n" + "="*50 + "\n")
            print(f"Lỗi khi xử lý sheet {sheet_name}: {e}")

# Xóa các sheet không có khác biệt khỏi file highlighted
sheets_to_remove = []
for sheet_name in wb_old_highlighted.sheetnames:
    if sheet_name not in sheets_with_differences:
        sheets_to_remove.append(sheet_name)
for sheet_name in sheets_to_remove:
    if sheet_name in wb_old_highlighted.sheetnames:
        wb_old_highlighted.remove(wb_old_highlighted[sheet_name])

sheets_to_remove = []
for sheet_name in wb_new_highlighted.sheetnames:
    if sheet_name not in sheets_with_differences:
        sheets_to_remove.append(sheet_name)
for sheet_name in sheets_to_remove:
    if sheet_name in wb_new_highlighted.sheetnames:
        wb_new_highlighted.remove(wb_new_highlighted[sheet_name])

# Đóng workbook gốc
wb_old.close()
wb_new.close()

# Lưu file Excel đã tô màu chỉ khi có sheet khác biệt
if sheets_with_differences:
    try:
        wb_old_highlighted.save(file_old_highlighted)
        wb_new_highlighted.save(file_new_highlighted)
        wb_old_highlighted.close()
        wb_new_highlighted.close()
        print(f"\nKết quả so sánh đã được lưu vào: {output_file}")
        print(f"File old đã tô màu: {file_old_highlighted} (chỉ chứa {len(sheets_with_differences)} sheet có khác biệt)")
        print(f"File new đã tô màu: {file_new_highlighted} (chỉ chứa {len(sheets_with_differences)} sheet có khác biệt)")
        print("Các ô có nội dung khác nhau đã được tô màu vàng.")
        print(f"Danh sách sheet có khác biệt:")
        for sheet in sheets_with_differences:
            print(f"  - {sheet}")
    except Exception as e:
        try:
            wb_old_highlighted.close()
            wb_new_highlighted.close()
        except:
            pass
        print(f"Lỗi khi lưu file tô màu: {e}")
        print(f"Kết quả so sánh đã được lưu vào: {output_file}")
        print(f"Có {len(sheets_with_differences)} sheet có khác biệt:")
        for sheet in sheets_with_differences:
            print(f"  - {sheet}")
else:
    wb_old_highlighted.close()
    wb_new_highlighted.close()
    print(f"\nKết quả so sánh đã được lưu vào: {output_file}")
    print("Tất cả sheet đều giống nhau. Không tạo file highlighted.")
    # Xóa file highlighted đã copy vì không cần thiết
    try:
        os.remove(file_old_highlighted)
        os.remove(file_new_highlighted)
    except:
        pass
