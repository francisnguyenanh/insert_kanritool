import os
import io
import json
import re
import zipfile
import tempfile
import datetime
import threading
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil
import pyodbc
import pandas as pd
from urllib.parse import quote_plus
from sqlalchemy import create_engine, text
from flask import Flask, render_template, request, jsonify, session, send_file
import genScriptFromExcel

import uuid
import subprocess

app = Flask(__name__)
app.secret_key = 'db_export_tool_secret_key_flask'
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024 * 1024


def get_script_dir():
    return os.path.dirname(os.path.abspath(__file__))


def get_main_conn():
    script_dir = get_script_dir()
    with open(os.path.join(script_dir, 'data_string.txt'), 'r', encoding='utf-8') as f:
        conn_str = f.read().strip()
    return pyodbc.connect(conn_str)


def get_file_conn():
    script_dir = get_script_dir()
    with open(os.path.join(script_dir, 'file_string.txt'), 'r', encoding='utf-8') as f:
        conn_str = f.read().strip()
    return pyodbc.connect(conn_str)


def export_data_file_helper(conn_file, lst_fileid, original_max_file_id):
    """Exports T_FILE_DATA and S_NUMBER_FILE – returns content as string."""
    file_id_keys = list(lst_fileid.keys())
    query = f"SELECT * FROM T_FILE_DATA WHERE FILE_ID IN ({', '.join(['?' for _ in file_id_keys])})"
    df = pd.read_sql(query, conn_file, params=file_id_keys)
    result = []
    if not df.empty:
        current_max_file_id = original_max_file_id
        insert_queries = []
        for _index, row in df.iterrows():
            values = []
            columns = []
            for col, value in row.items():
                if col == 'TIME_STAMP':
                    continue
                columns.append(col)
                if pd.isnull(value):
                    values.append('NULL')
                elif isinstance(value, bool):
                    values.append('1' if value else '0')
                elif col == 'FILE_ID':
                    current_max_file_id += 1
                    values.append(f"'{lst_fileid.get(value)}'")
                elif isinstance(value, (int, float)):
                    values.append(f"'{int(value)}'")
                elif isinstance(value, pd.Timestamp):
                    values.append(f"'{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                elif isinstance(value, str):
                    values.append(f"N'{value}'")
                elif isinstance(value, bytes):
                    values.append(f"0x{value.hex()}")
                else:
                    values.append(f"'{str(value)}'")
            insert_queries.append(
                f"INSERT INTO T_FILE_DATA ({', '.join(columns)}) VALUES ({', '.join(values)});"
            )
        result.extend(insert_queries)
        # S_NUMBER_FILE
        df2 = pd.read_sql("SELECT * FROM S_NUMBER_FILE", conn_file)
        set_clause = []
        for col, value in df2.iloc[0].items():
            if col in ('CREATE_USER', 'CREATE_DATE', 'CREATE_PC', 'TIME_STAMP', 'DELETE_FLG'):
                continue
            if isinstance(value, str):
                set_clause.append(f"{col} = N'{value}'")
            elif ('ID' in col or 'KBN' in col or 'CD' in col) and isinstance(value, (int, float)):
                set_clause.append(f"{col} = {int(value)}")
            elif col == 'CURRENT_NUMBER':
                set_clause.append(f"{col} = {current_max_file_id}")
            elif pd.isnull(value):
                set_clause.append(f"{col} = NULL")
            elif isinstance(value, pd.Timestamp):
                set_clause.append(f"{col} = '{value.strftime('%Y-%m-%d %H:%M:%S')}'")
            else:
                set_clause.append(f"{col} = {repr(value)}")
        result.append(f"UPDATE S_NUMBER_FILE SET {', '.join(set_clause)};")
    return '\n'.join(result)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/connect', methods=['POST'])
def connect_to_db():
    data = request.get_json()
    old_system_id = (data.get('old_system_id') or '').strip()

    if not old_system_id:
        return jsonify({'status': 'error', 'message': 'Please enter the Old System ID before connecting to the database.'})

    try:
        script_dir = get_script_dir()
        with open(os.path.join(script_dir, 'table_logic.txt'), 'r', encoding='utf-8') as f:
            table_names = [line.strip() for line in f if line.strip()]

        conn = get_main_conn()
        cursor = conn.cursor()

        matching_tables = []
        for table_name in table_names:
            cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE SYSTEM_ID = ?", old_system_id)
            if cursor.fetchone()[0] > 0:
                matching_tables.append(table_name)

        conn.close()
        matching_tables = sorted(matching_tables)
        is_need_fileID = 'T_FILE_LINK_KIHON_PJ_GAMEN' in matching_tables

        session['matching_tables'] = matching_tables
        session['is_need_fileID'] = is_need_fileID

        if not matching_tables:
            return jsonify({'status': 'warning', 'message': 'No data found for the given System ID in any table.'})

        return jsonify({
            'status': 'success',
            'message': 'Loaded matching table names successfully!',
            'matching_tables': matching_tables,
            'is_need_fileID': is_need_fileID,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Failed to load table names: {e}'})


@app.route('/export_multi', methods=['POST'])
def export_data_multi():
    data = request.get_json()
    old_system_id = (data.get('old_system_id') or '').strip()
    new_system_id = (data.get('new_system_id') or '').strip()
    current_max_file_id_str = (data.get('current_max_file_id') or '').strip()
    matching_tables = data.get('matching_tables', [])
    is_need_fileID = data.get('is_need_fileID', False)

    if is_need_fileID:
        if not current_max_file_id_str:
            return jsonify({'status': 'error', 'message': 'Please enter current max file ID.'})
        current_max_file_id = int(current_max_file_id_str)
    else:
        current_max_file_id = 0

    original_max_file_id = current_max_file_id

    if not old_system_id or not new_system_id:
        return jsonify({'status': 'error', 'message': 'Please enter both Old System ID and New System ID.'})
    if not matching_tables:
        return jsonify({'status': 'error', 'message': 'No tables found to export.'})

    columns_to_convert = ['ZENKAKU_MOJI_SU', 'HANKAKU_MOJI_SU', 'SEISU_KETA', 'SYOUSU_KETA']

    try:
        conn = get_main_conn()
        lst_fileid = {}
        have_file = False
        all_files = {}

        for table_name in matching_tables:
            if not table_name.strip():
                continue

            df = pd.read_sql(f"SELECT * FROM {table_name} WHERE SYSTEM_ID = ?", conn, params=[old_system_id])
            if df.empty:
                continue

            insert_queries = []

            if table_name == 'T_KIHON_PJ':
                set_clause = []
                for col, value in df.iloc[0].items():
                    if col in ('SYSTEM_ID', 'TIME_STAMP', 'DELETE_FLG', 'KOUSHIN_FUKA_FLG'):
                        continue
                    if isinstance(value, str):
                        set_clause.append(f"{col} = N'{value}'")
                    elif ('ID' in col or 'KBN' in col or 'CD' in col or 'SEQ' in col) and isinstance(value, (int, float)):
                        set_clause.append(f"{col} = {int(value)}")
                    elif pd.isnull(value):
                        set_clause.append(f"{col} = NULL")
                    elif isinstance(value, pd.Timestamp):
                        set_clause.append(f"{col} = '{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                    else:
                        set_clause.append(f"{col} = {repr(value)}")
                insert_queries.append(
                    f"UPDATE {table_name} SET {', '.join(set_clause)} WHERE SYSTEM_ID = '{new_system_id}';"
                )
            else:
                insert_queries.append(f"DELETE FROM {table_name} WHERE SYSTEM_ID = '{new_system_id}';")
                for _index, row in df.iterrows():
                    values = []
                    columns = []
                    for col, value in row.items():
                        if col == 'TIME_STAMP':
                            continue
                        columns.append(col)
                        if pd.isnull(value):
                            values.append('NULL')
                        elif col == 'SYSTEM_ID':
                            values.append(f"'{new_system_id}'")
                        elif table_name == 'T_FILE_LINK_KIHON_PJ_GAMEN' and col == 'FILE_ID':
                            have_file = True
                            current_max_file_id += 1
                            lst_fileid[value] = current_max_file_id
                            values.append(f"'{current_max_file_id}'")
                        elif table_name == 'T_FILE_LINK_KIHON_PJ_GAMEN' and isinstance(value, pd.Timestamp):
                            values.append(f"'{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                        elif ('ID' in col or 'KBN' in col or 'CD' in col or 'SEQ' in col) and isinstance(value, (int, float)):
                            values.append(f"'{int(value)}'")
                        elif col in columns_to_convert and isinstance(value, (int, float)):
                            values.append(f"'{int(value)}'")
                        elif isinstance(value, str):
                            values.append(f"N'{value}'")
                        else:
                            values.append(f"'{str(value)}'")
                    insert_queries.append(
                        f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(values)});"
                    )
            all_files[table_name + '.sql'] = '\n'.join(insert_queries)

        conn.close()

        if have_file:
            conn_file = get_file_conn()
            filedata = export_data_file_helper(conn_file, lst_fileid, original_max_file_id)
            all_files['T_FILE_DATA_and_S_NUMBER_FILE.sql'] = filedata
            conn_file.close()

        # Create a zip file in memory
        import io, zipfile
        mem_zip = io.BytesIO()
        with zipfile.ZipFile(mem_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            for fname, content in all_files.items():
                zf.writestr(fname, content)
        mem_zip.seek(0)
        return send_file(mem_zip, as_attachment=True, download_name='exported_sql_files.zip', mimetype='application/zip')
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'})


@app.route('/export_single', methods=['POST'])
def export_data_single():
    data = request.get_json()
    old_system_id = (data.get('old_system_id') or '').strip()
    new_system_id = (data.get('new_system_id') or '').strip()
    current_max_file_id_str = (data.get('current_max_file_id') or '').strip()
    matching_tables = data.get('matching_tables', [])
    is_need_fileID = data.get('is_need_fileID', False)

    if is_need_fileID:
        if not current_max_file_id_str:
            return jsonify({'status': 'error', 'message': 'Please enter current max file ID.'})
        current_max_file_id = int(current_max_file_id_str)
    else:
        current_max_file_id = 0

    original_max_file_id = current_max_file_id

    if not old_system_id or not new_system_id:
        return jsonify({'status': 'error', 'message': 'Please enter both Old System ID and New System ID.'})
    if not matching_tables:
        return jsonify({'status': 'error', 'message': 'No tables found to export.'})

    columns_to_convert = ['ZENKAKU_MOJI_SU', 'HANKAKU_MOJI_SU', 'SEISU_KETA', 'SYOUSU_KETA']

    try:
        conn = get_main_conn()
        all_queries = []
        lst_fileid = {}
        have_file = False

        for table_name in matching_tables:
            if not table_name.strip():
                continue

            df = pd.read_sql(f"SELECT * FROM {table_name} WHERE SYSTEM_ID = ?", conn, params=[old_system_id])
            if df.empty:
                continue

            insert_queries = []

            if table_name == 'T_KIHON_PJ':
                set_clause = []
                for col, value in df.iloc[0].items():
                    if col in ('SYSTEM_ID', 'TIME_STAMP', 'DELETE_FLG', 'KOUSHIN_FUKA_FLG'):
                        continue
                    if isinstance(value, str):
                        set_clause.append(f"{col} = N'{value}'")
                    elif ('ID' in col or 'KBN' in col or 'CD' in col or 'SEQ' in col) and isinstance(value, (int, float)):
                        set_clause.append(f"{col} = {int(value)}")
                    elif pd.isnull(value):
                        set_clause.append(f"{col} = NULL")
                    elif isinstance(value, pd.Timestamp):
                        set_clause.append(f"{col} = '{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                    else:
                        set_clause.append(f"{col} = {repr(value)}")
                insert_queries.append(
                    f"UPDATE {table_name} SET {', '.join(set_clause)} WHERE SYSTEM_ID = '{new_system_id}';"
                )
            else:
                insert_queries.append(f"DELETE FROM {table_name} WHERE SYSTEM_ID = '{new_system_id}';")
                for _index, row in df.iterrows():
                    values = []
                    columns = []
                    for col, value in row.items():
                        if col == 'TIME_STAMP':
                            continue
                        columns.append(col)
                        if pd.isnull(value):
                            values.append('NULL')
                        elif col == 'SYSTEM_ID':
                            values.append(f"'{new_system_id}'")
                        elif table_name == 'T_FILE_LINK_KIHON_PJ_GAMEN' and col == 'FILE_ID':
                            have_file = True
                            current_max_file_id += 1
                            lst_fileid[value] = current_max_file_id
                            values.append(f"'{current_max_file_id}'")
                        elif table_name == 'T_FILE_LINK_KIHON_PJ_GAMEN' and isinstance(value, pd.Timestamp):
                            values.append(f"'{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                        elif ('ID' in col or 'KBN' in col or 'CD' in col or 'SEQ' in col) and isinstance(value, (int, float)):
                            values.append(f"'{int(value)}'")
                        elif col in columns_to_convert and isinstance(value, (int, float)):
                            values.append(f"'{int(value)}'")
                        elif isinstance(value, str):
                            values.append(f"N'{value}'")
                        else:
                            values.append(f"'{str(value)}'")
                    insert_queries.append(
                        f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(values)});"
                    )
            all_queries.extend(insert_queries)
            all_queries.append("GO")

        conn.close()

        if have_file:
            conn_file = get_file_conn()
            filedata = export_data_file_helper(conn_file, lst_fileid, original_max_file_id)
            all_queries.append(filedata)
            conn_file.close()

        sql_bytes = io.BytesIO()
        sql_bytes.write('\n'.join(all_queries).encode('utf-8'))
        sql_bytes.seek(0)
        return send_file(sql_bytes, as_attachment=True, download_name='all_tables.sql', mimetype='text/plain')
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'})



@app.route('/compare_excel', methods=['POST'])
def compare_excel():
    if 'file_old' not in request.files or 'file_new' not in request.files:
        return jsonify({'status': 'error', 'message': 'Vui lòng chọn cả hai file.'})
    
    file_old = request.files['file_old']
    file_new = request.files['file_new']
    
    if not file_old.filename or not file_new.filename:
        return jsonify({'status': 'error', 'message': 'Chưa chọn đủ file.'})

    tmp_dir = tempfile.mkdtemp()
    try:
        path_old = os.path.join(tmp_dir, 'old.xlsx')
        path_new = os.path.join(tmp_dir, 'new.xlsx')
        file_old.save(path_old)
        file_new.save(path_new)
        
        output_file = os.path.join(tmp_dir, 'comparison_result.txt')
        file_old_highlighted = os.path.join(tmp_dir, 'old_highlighted.xlsx')
        file_new_highlighted = os.path.join(tmp_dir, 'new_highlighted.xlsx')
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        shutil.copy2(path_old, file_old_highlighted)
        shutil.copy2(path_new, file_new_highlighted)
        
        wb_old = load_workbook(path_old, data_only=True)
        wb_new = load_workbook(path_new, data_only=True)
        wb_old_highlighted = load_workbook(file_old_highlighted)
        wb_new_highlighted = load_workbook(file_new_highlighted)
        
        sheets_with_differences = []
        
        with open(output_file, 'w', encoding='utf-8') as f:
            sheets_old = wb_old.sheetnames
            sheets_new = wb_new.sheetnames
            if len(sheets_old) != len(sheets_new):
                f.write("Số lượng sheet không giống nhau!\n")
                f.write(f"File old có {len(sheets_old)} sheets: {', '.join(sheets_old)}\n")
                f.write(f"File new có {len(sheets_new)} sheets: {', '.join(sheets_new)}\n\n")
            else:
                f.write("Số lượng sheet giống nhau. Kết quả so sánh:\n\n")
            
            common_sheets = set(sheets_old) & set(sheets_new)
            only_in_old = set(sheets_old) - set(sheets_new)
            only_in_new = set(sheets_new) - set(sheets_old)
            
            if only_in_old:
                f.write(f"Sheet chỉ có trong file old: {', '.join(only_in_old)}\n")
            if only_in_new:
                f.write(f"Sheet chỉ có trong file new: {', '.join(only_in_new)}\n")
            if only_in_old or only_in_new:
                f.write("\n")
                
            for sheet_name in common_sheets:
                try:
                    ws_old = wb_old[sheet_name]
                    ws_new = wb_new[sheet_name]
                    ws_old_highlighted = wb_old_highlighted[sheet_name]
                    ws_new_highlighted = wb_new_highlighted[sheet_name]
                    
                    max_row = max(ws_old.max_row, ws_new.max_row)
                    max_col = max(ws_old.max_column, ws_new.max_column)
                    different_cells = 0
                    different_positions = []
                    has_differences = False
                    
                    for row in range(1, max_row + 1):
                        for col in range(1, max_col + 1):
                            value_old = ws_old.cell(row=row, column=col).value
                            value_new = ws_new.cell(row=row, column=col).value
                            clean_value_old = str(value_old).strip() if value_old is not None else ''
                            clean_value_new = str(value_new).strip() if value_new is not None else ''
                            if clean_value_old != clean_value_new:
                                different_cells += 1
                                has_differences = True
                                try:
                                    cell_old = ws_old_highlighted.cell(row=row, column=col)
                                    merged_old = False
                                    for merged_range in ws_old_highlighted.merged_cells.ranges:
                                        if cell_old.coordinate in merged_range:
                                            for merged_cell in merged_range.cells:
                                                ws_old_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                            merged_old = True
                                            break
                                    if not merged_old:
                                        cell_old.fill = yellow_fill
                                        
                                    cell_new = ws_new_highlighted.cell(row=row, column=col)
                                    merged_new = False
                                    for merged_range in ws_new_highlighted.merged_cells.ranges:
                                        if cell_new.coordinate in merged_range:
                                            for merged_cell in merged_range.cells:
                                                ws_new_highlighted.cell(row=merged_cell[0], column=merged_cell[1]).fill = yellow_fill
                                            merged_new = True
                                            break
                                    if not merged_new:
                                        cell_new.fill = yellow_fill
                                except Exception as e:
                                    pass
                                
                                if len(different_positions) < 20:
                                    col_letter = ws_old.cell(row=row, column=col).column_letter
                                    different_positions.append({
                                        'position': f"{col_letter}{row}",
                                        'value_old': clean_value_old if clean_value_old != '' else "NULL",
                                        'value_new': clean_value_new if clean_value_new != '' else "NULL"
                                    })
                                    
                    values_old = set()
                    values_new = set()
                    non_empty_rows_old = 0
                    non_empty_rows_new = 0
                    
                    for row in range(1, ws_old.max_row + 1):
                        has_data = False
                        for col in range(1, ws_old.max_column + 1):
                            value = ws_old.cell(row=row, column=col).value
                            if value is not None and str(value).strip() != '':
                                clean_value = str(value).strip()
                                values_old.add(clean_value)
                                has_data = True
                        if has_data:
                            non_empty_rows_old += 1
                            
                    for row in range(1, ws_new.max_row + 1):
                        has_data = False
                        for col in range(1, ws_new.max_column + 1):
                            value = ws_new.cell(row=row, column=col).value
                            if value is not None and str(value).strip() != '':
                                clean_value = str(value).strip()
                                values_new.add(clean_value)
                                has_data = True
                        if has_data:
                            non_empty_rows_new += 1
                            
                    if has_differences or values_old != values_new:
                        sheets_with_differences.append(sheet_name)
                        f.write(f"Sheet: {sheet_name}\n")
                        f.write(f"Số dòng có dữ liệu trong file old: {non_empty_rows_old}\n")
                        f.write(f"Số dòng có dữ liệu trong file new: {non_empty_rows_new}\n")
                        f.write(f"Tổng số giá trị khác rỗng trong file old: {len(values_old)}\n")
                        f.write(f"Tổng số giá trị khác rỗng trong file new: {len(values_new)}\n")
                        f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                        if different_positions:
                            f.write(f"Ví dụ {min(len(different_positions), 20)} vị trí khác nhau đầu tiên:\n")
                            for pos in different_positions:
                                f.write(f"  {pos['position']}: old='{pos['value_old']}' vs new='{pos['value_new']}'\n")
                        
                        only_in_old_vals = values_old - values_new
                        only_in_new_vals = values_new - values_old
                        
                        if only_in_old_vals:
                            f.write(f"Giá trị chỉ có trong file old ({len(only_in_old_vals)} giá trị):\n")
                            for value in sorted(list(only_in_old_vals)[:20]):
                                f.write(f"  {value}\n")
                            if len(only_in_old_vals) > 20:
                                f.write(f"  ... và {len(only_in_old_vals) - 20} giá trị khác\n")
                        if only_in_new_vals:
                            f.write(f"Giá trị chỉ có trong file new ({len(only_in_new_vals)} giá trị):\n")
                            for value in sorted(list(only_in_new_vals)[:20]):
                                f.write(f"  {value}\n")
                            if len(only_in_new_vals) > 20:
                                f.write(f"  ... và {len(only_in_new_vals) - 20} giá trị khác\n")
                        f.write("\n" + "="*50 + "\n")
                    else:
                        f.write(f"Sheet: {sheet_name}\n")
                        f.write("Nội dung giống hệt nhau (bỏ qua dòng trắng và cấu trúc).\n")
                        f.write(f"Số ô có nội dung khác nhau: {different_cells}\n")
                        f.write("\n" + "="*50 + "\n")
                except Exception as e:
                    f.write(f"Sheet: {sheet_name}\nLỗi khi xử lý sheet: {e}\n" + "="*50 + "\n")
                    
        sheets_to_remove = [sn for sn in wb_old_highlighted.sheetnames if sn not in sheets_with_differences]
        for sn in sheets_to_remove:
            if sn in wb_old_highlighted.sheetnames:
                wb_old_highlighted.remove(wb_old_highlighted[sn])
                
        sheets_to_remove = [sn for sn in wb_new_highlighted.sheetnames if sn not in sheets_with_differences]
        for sn in sheets_to_remove:
            if sn in wb_new_highlighted.sheetnames:
                wb_new_highlighted.remove(wb_new_highlighted[sn])

        wb_old.close()
        wb_new.close()
        
        mem_zip = io.BytesIO()
        import zipfile
        with zipfile.ZipFile(mem_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            zf.write(output_file, 'comparison_result.txt')
            if sheets_with_differences:
                wb_old_highlighted.save(file_old_highlighted)
                wb_new_highlighted.save(file_new_highlighted)
                zf.write(file_old_highlighted, 'old_highlighted.xlsx')
                zf.write(file_new_highlighted, 'new_highlighted.xlsx')
                
        wb_old_highlighted.close()
        wb_new_highlighted.close()
        
        mem_zip.seek(0)
        return send_file(mem_zip, as_attachment=True, download_name='comparison_results.zip', mimetype='application/zip')
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Lỗi khi xử lý: {str(e)}'})
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

# ──────────────────────────────────────────────
#  Video Converter API
# ──────────────────────────────────────────────

@app.route('/convert_video', methods=['POST'])
def convert_video():
    if 'video' not in request.files:
        return jsonify({'status': 'error', 'message': 'Không tìm thấy file video.'})
    
    file = request.files['video']
    if not file.filename:
        return jsonify({'status': 'error', 'message': 'Chưa chọn file.'})
    
    tmp_dir = tempfile.mkdtemp()
    try:
        unique_id = str(uuid.uuid4())
        ext = os.path.splitext(file.filename)[1]
        if not ext:
            ext = '.mp4'
        input_path = os.path.join(tmp_dir, f"input_{unique_id}{ext}")
        output_path = os.path.join(tmp_dir, f"output_{unique_id}.mp3")
        
        file.save(input_path)
        
        # Call FFMPEG
        cmd = [
            'ffmpeg', '-i', input_path,
            '-q:a', '0', '-map', 'a',
            output_path, '-y'
        ]
        
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        # Send file back
        with open(output_path, 'rb') as f:
            mp3_data = io.BytesIO(f.read())
            
        mp3_data.seek(0)
        
        download_name = os.path.splitext(file.filename)[0] + '.mp3'
        return send_file(mp3_data, as_attachment=True, download_name=download_name, mimetype='audio/mpeg')
        
    except subprocess.CalledProcessError as e:
        return jsonify({'status': 'error', 'message': f'Lỗi khi convert video (FFMPEG error).'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Lỗi hệ thống: {str(e)}'})
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

# ──────────────────────────────────────────────
#  Config API
# ──────────────────────────────────────────────

CONFIG_FILES = {
    'data_string':       'data_string.txt',
    'file_string':       'file_string.txt',
    'dataconnectstring': 'dataconnectstring.txt',
    'table':             'table.txt',
    'table_logic':       'table_logic.txt',
}


@app.route('/get_config', methods=['GET'])
def get_config():
    script_dir = get_script_dir()
    result = {}
    for key, filename in CONFIG_FILES.items():
        path = os.path.join(script_dir, filename)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                result[key] = f.read()
        except Exception:
            result[key] = ''
    return jsonify({'status': 'success', 'config': result})


@app.route('/save_config', methods=['POST'])
def save_config():
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'message': 'No data provided.'})
    script_dir = get_script_dir()
    try:
        for key, filename in CONFIG_FILES.items():
            if key in data:
                path = os.path.join(script_dir, filename)
                with open(path, 'w', encoding='utf-8') as f:
                    f.write(data[key])
        return jsonify({'status': 'success', 'message': 'Configuration saved successfully.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Failed to save configuration: {e}'})


# ──────────────────────────────────────────────
#  Tab 2: Gen Script from Excel
# ──────────────────────────────────────────────


# ──── Gen Script Config API ────

EXCEL_CONFIG_FILES = {
    'username_id': 'usernameID.txt',
    'table_info':  'TABLE_INFO.txt',
}


@app.route('/get_excel_config', methods=['GET'])
def get_excel_config():
    script_dir = get_script_dir()
    result = {}
    # Plain text files
    for key, filename in EXCEL_CONFIG_FILES.items():
        path = os.path.join(script_dir, filename)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                result[key] = f.read()
        except Exception:
            result[key] = ''
    # genscript_config.json
    config_path = os.path.join(script_dir, 'genscript_config.json')
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            result['genscript_config'] = json.load(f)
    except Exception:
        result['genscript_config'] = {}
    return jsonify({'status': 'success', 'config': result})


@app.route('/save_excel_config', methods=['POST'])
def save_excel_config():
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'message': 'No data provided.'})
    script_dir = get_script_dir()
    try:
        # Save usernameID.txt
        if 'username_id' in data:
            with open(os.path.join(script_dir, 'usernameID.txt'), 'w', encoding='utf-8') as f:
                f.write(str(data['username_id']).strip())
        # Save TABLE_INFO.txt — validate JSON first
        if 'table_info' in data:
            try:
                json.loads(data['table_info'])  # validate
            except Exception as e:
                return jsonify({'status': 'error', 'message': f'TABLE_INFO.txt: invalid JSON — {e}'})
            with open(os.path.join(script_dir, 'TABLE_INFO.txt'), 'w', encoding='utf-8') as f:
                f.write(data['table_info'])
        # Save genscript_config.json
        if 'genscript_config' in data:
            with open(os.path.join(script_dir, 'genscript_config.json'), 'w', encoding='utf-8') as f:
                json.dump(data['genscript_config'], f, ensure_ascii=False, indent=2)
        # Reload constants in the genScriptFromExcel module
        genScriptFromExcel.reload_config()
        # Reset username ID counter so it re-reads the updated file
        genScriptFromExcel._username_id_counter = None
        return jsonify({'status': 'success', 'message': 'Configuration saved successfully.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Failed to save configuration: {e}'})


import base64


def validate_insert_columns(sql_content):
    """Parse INSERT statements, compare columns against DB, return list of warning strings."""
    import re

    # Load skip_check_columns from genscript_config.json
    skip_cols = set()
    try:
        config_path = os.path.join(get_script_dir(), 'genscript_config.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        skip_cols = {c.upper() for c in cfg.get('skip_check_columns', [])}
    except Exception:
        pass

    # Collect unique set of columns per table from INSERT statements
    table_cols = {}  # {table_name: set_of_columns}
    pattern = re.compile(
        r'INSERT\s+INTO\s+(\w+)\s*\(([^)]+)\)',
        re.IGNORECASE
    )
    for m in pattern.finditer(sql_content):
        tname = m.group(1).upper()
        cols = {c.strip().upper() for c in m.group(2).split(',')}
        if tname not in table_cols:
            table_cols[tname] = cols
        else:
            table_cols[tname] |= cols  # union: accumulate all cols seen

    if not table_cols:
        return []

    try:
        conn = get_main_conn()
    except Exception as e:
        return [f'[DB connect failed — column check skipped] {e}']

    warnings = []
    try:
        cursor = conn.cursor()
        for tname, insert_cols in sorted(table_cols.items()):
            try:
                cursor.execute(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_NAME = ? ORDER BY ORDINAL_POSITION",
                    tname
                )
                db_cols = {row[0].upper() for row in cursor.fetchall()} - skip_cols
                if not db_cols:
                    warnings.append(f"⚠️ Table {tname}: not found in DB (skipped).")
                    continue
                effective_insert = insert_cols - skip_cols
                missing = sorted(db_cols - effective_insert)
                extra   = sorted(effective_insert - db_cols)
                if missing and extra:
                    warnings.append(
                        f"⚠️ Table {tname}: INSERT に欠けているカラム: {', '.join(missing)} / "
                        f"余分なカラム: {', '.join(extra)}"
                    )
                elif missing:
                    warnings.append(
                        f"⚠️ Table {tname}: INSERT に欠けているカラム: {', '.join(missing)}"
                    )
                elif extra:
                    warnings.append(
                        f"⚠️ Table {tname}: INSERT にテーブル内にないカラム: {', '.join(extra)}"
                    )
                else:
                    warnings.append(f"✅ Table {tname}: OK (全カラム一致)")
            except Exception as e:
                warnings.append(f"⚠️ Table {tname}: check failed — {e}")
    finally:
        conn.close()

    return warnings


@app.route('/gen_excel', methods=['POST'])
def gen_excel():
    if 'excel_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No Excel file uploaded.'})

    excel_file = request.files['excel_file']
    if not excel_file.filename:
        return jsonify({'status': 'error', 'message': 'No Excel file selected.'})

    system_id   = request.form.get('system_id', '').strip()
    system_date = request.form.get('system_date', '').strip()

    # Save uploaded Excel to a temp file
    tmp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        excel_file.save(tmp_excel.name)
        tmp_excel.close()
    except Exception:
        tmp_excel.close()
        try:
            os.unlink(tmp_excel.name)
        except OSError:
            pass
        return jsonify({'status': 'error', 'message': 'Failed to save uploaded file.'})

    tmp_output = tempfile.NamedTemporaryFile(delete=False, suffix='.sql', mode='w', encoding='utf-8')
    tmp_output.close()

    try:
        # Reload config before each run so any saved changes take effect
        genScriptFromExcel.reload_config()
        # Override appX globals per-request
        now = datetime.datetime.now()
        genScriptFromExcel.systemid_value    = system_id  if system_id   else f"{now.hour:02d}{now.minute:02d}{now.second:02d}"
        genScriptFromExcel.system_date_value = system_date if system_date else now.strftime('%Y-%m-%d')

        # Path to TABLE_INFO.txt stored next to app.py
        table_info_path = os.path.join(get_script_dir(), 'TABLE_INFO.txt')

        genScriptFromExcel.all_tables_in_sequence(tmp_excel.name, table_info_path, tmp_output.name)

        with open(tmp_output.name, 'r', encoding='utf-8') as f:
            sql_content = f.read()

        # Validate columns against DB
        warnings = validate_insert_columns(sql_content)

        sql_b64 = base64.b64encode(sql_content.encode('utf-8')).decode('ascii')
        return jsonify({
            'status': 'success',
            'sql_b64': sql_b64,
            'warnings': warnings,
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'An error occurred: {str(e)}'})
    finally:
        for path in (tmp_excel.name, tmp_output.name):
            try:
                os.unlink(path)
            except OSError:
                pass


# ─────────────────────────────────────────────────────────────
#  Reproduce DB — config helpers
# ─────────────────────────────────────────────────────────────

_REPRODUCE_CONFIG_FILE          = 'reproduce_config.json'
_SELECTED_TABLES_FILE          = 'reproduce_selected_tables.json'
_REPRODUCE_CONFIG_DEFAULTS = {
    'src_conn_str':   '',
    'dst_conn_str':   '',
    'src_db_name':    '',
    'dst_db_name':    '',
    'rows_per_table': 1000,
}


def _get_reproduce_config():
    path = os.path.join(get_script_dir(), _REPRODUCE_CONFIG_FILE)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {**_REPRODUCE_CONFIG_DEFAULTS, **data}
    except Exception:
        return dict(_REPRODUCE_CONFIG_DEFAULTS)


def _save_reproduce_config(data):
    path = os.path.join(get_script_dir(), _REPRODUCE_CONFIG_FILE)
    cfg = {k: data.get(k, v) for k, v in _REPRODUCE_CONFIG_DEFAULTS.items()}
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return cfg


# ─────────────────────────────────────────────────────────────
#  Reproduce DB — background migration job state
# ─────────────────────────────────────────────────────────────

_migration_lock = threading.Lock()
_migration_state = {
    'running':  False,
    'done':     False,
    'success':  False,
    'progress': 0,
    'label':    '',
    'logs':     [],
}


def _run_migration(cfg):
    """Background thread: copies tables from a remote SQL Server to a new local DB."""
    global _migration_state

    def log(msg, level='INFO'):
        prefix = {
            'INFO':    '[INFO]',
            'ERROR':   '[ERROR]',
            'WARN':    '[WARN]',
            'SUCCESS': '[SUCCESS]',
        }.get(level, '[INFO]')
        with _migration_lock:
            _migration_state['logs'].append(f"{prefix} {msg}")

    def set_progress(pct, label=''):
        with _migration_lock:
            _migration_state['progress'] = pct
            if label:
                _migration_state['label'] = label

    def finish(success):
        with _migration_lock:
            _migration_state['done']    = True
            _migration_state['running'] = False
            _migration_state['success'] = success

    try:
        src_conn_str = (cfg.get('src_conn_str') or '').strip()
        dst_conn_str = (cfg.get('dst_conn_str') or '').strip()
        src_db_name  = (cfg.get('src_db_name')  or '').strip()
        dst_db_name  = (cfg.get('dst_db_name')  or '').strip()
        rows         = int(cfg.get('rows_per_table') or 1000)

        if not src_conn_str or not dst_conn_str or not src_db_name or not dst_db_name:
            log('Missing required configuration (connection strings or DB names).', 'ERROR')
            finish(False)
            return

        # Validate destination DB name to prevent injection in CREATE DATABASE
        if not re.match(r'^[A-Za-z0-9_\-]+$', dst_db_name):
            log(
                f"Invalid destination DB name '{dst_db_name}'. "
                "Use letters, digits, underscores, or hyphens only.", 'ERROR'
            )
            finish(False)
            return

        # ── Step 1: Create destination DB if not exists ──
        log(f"Connecting to destination server to create database '{dst_db_name}'…")
        set_progress(2, 'Creating destination database…')
        try:
            conn_master = pyodbc.connect(dst_conn_str, autocommit=True)
            cur = conn_master.cursor()
            cur.execute("SELECT COUNT(*) FROM sys.databases WHERE name = ?", dst_db_name)
            if cur.fetchone()[0] == 0:
                cur.execute(f"CREATE DATABASE [{dst_db_name}]")
                log(f"Database '{dst_db_name}' created.", 'SUCCESS')
            else:
                log(f"Database '{dst_db_name}' already exists — skipping creation.")
            conn_master.close()
        except Exception as e:
            log(f"Failed to create destination database: {e}", 'ERROR')
            finish(False)
            return

        # ── Step 2: Connect to source DB (and optionally fetch table list) ──
        log(f"Connecting to source database '{src_db_name}'…")
        set_progress(5, 'Connecting to source…')
        try:
            upper = src_conn_str.upper()
            if 'DATABASE=' not in upper and 'INITIAL CATALOG=' not in upper:
                src_full = src_conn_str + f";DATABASE={src_db_name}"
            else:
                src_full = src_conn_str
            source_conn = pyodbc.connect(src_full)

            selected_tables = cfg.get('selected_tables')
            if selected_tables:
                table_list = list(selected_tables)
                log(f"{len(table_list)} bảng được chọn để copy.")
            else:
                tables_df = pd.read_sql(
                    "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
                    "WHERE TABLE_TYPE = 'BASE TABLE' ORDER BY TABLE_NAME",
                    source_conn
                )
                table_list = tables_df['TABLE_NAME'].tolist()
                log(f"Found {len(table_list)} table(s) to copy.")
        except Exception as e:
            log(f"Failed to connect to source database: {e}", 'ERROR')
            finish(False)
            return

        # ── Step 3: Open a raw pyodbc connection to the destination DB ──
        # (Avoids SQLAlchemy's has_table() which generates CAST(? AS NVARCHAR(max))
        #  — incompatible with the legacy [Microsoft][ODBC SQL Server Driver] → HY104)
        try:
            dst_full = dst_conn_str + f";DATABASE={dst_db_name}"
            dest_raw = pyodbc.connect(dst_full, autocommit=False)
        except Exception as e:
            log(f"Failed to connect to destination database: {e}", 'ERROR')
            source_conn.close()
            finish(False)
            return

        # ── Helpers for raw-pyodbc writes ────────────────────────────────────

        def _pandas_dtype_to_sql(dtype):
            """Map a pandas dtype to a SQL Server column type string."""
            from pandas.api.types import (
                is_bool_dtype, is_integer_dtype,
                is_float_dtype, is_datetime64_any_dtype,
            )
            if is_bool_dtype(dtype):             return "BIT"
            if is_integer_dtype(dtype):          return "BIGINT"
            if is_float_dtype(dtype):            return "FLOAT"
            if is_datetime64_any_dtype(dtype):   return "DATETIME2"
            return "NVARCHAR(MAX)"

        def _safe_val(v):
            """Convert NaN / NaT / None to None; leave everything else intact."""
            if v is None:
                return None
            try:
                return None if pd.isnull(v) else v
            except (TypeError, ValueError):
                return v

        # ── Step 4: Copy each table ──
        total         = len(table_list)
        success_count = 0
        error_count   = 0

        for i, table in enumerate(table_list):
            pct = 5 + int((i / total) * 90) if total else 95
            set_progress(pct, f"Đang copy bảng {i + 1}/{total}: {table}…")
            log(f"({i + 1}/{total}) Đang copy bảng [{table}]…")
            try:
                if rows > 0:
                    query = f"SELECT TOP {rows} * FROM [{table}] ORDER BY 1 DESC"
                else:
                    query = f"SELECT * FROM [{table}]"
                df = pd.read_sql(query, source_conn)
                row_count = len(df)

                cur = dest_raw.cursor()

                # Drop existing table — uses an identifier in the SQL text
                # (no driver-parameter binding) to avoid the HY104 issue.
                cur.execute(
                    f"IF OBJECT_ID(N'[{table}]', N'U') IS NOT NULL "
                    f"DROP TABLE [{table}]"
                )

                # Build CREATE TABLE from DataFrame dtypes
                col_defs = ", ".join(
                    f"[{col}] {_pandas_dtype_to_sql(dtype)} NULL"
                    for col, dtype in df.dtypes.items()
                )
                cur.execute(f"CREATE TABLE [{table}] ({col_defs})")

                if row_count > 0:
                    col_names   = ", ".join(f"[{c}]" for c in df.columns)
                    placeholders = ", ".join("?" for _ in df.columns)
                    insert_sql  = (
                        f"INSERT INTO [{table}] ({col_names}) "
                        f"VALUES ({placeholders})"
                    )
                    # Convert NA → None for pyodbc compatibility
                    data_rows = [
                        tuple(_safe_val(v) for v in row)
                        for row in df.itertuples(index=False, name=None)
                    ]
                    # Insert in chunks of 500 rows to stay within driver limits
                    chunk = 500
                    for start in range(0, row_count, chunk):
                        cur.executemany(insert_sql, data_rows[start:start + chunk])

                dest_raw.commit()
                log(f"Đang copy bảng [{table}]… Thành công ({row_count} dòng)", 'SUCCESS')
                success_count += 1
            except Exception as e:
                dest_raw.rollback()
                log(f"Bảng [{table}] thất bại: {e}", 'ERROR')
                error_count += 1

        source_conn.close()
        dest_raw.close()
        set_progress(100, 'Done')

        if error_count == 0:
            log(f"Migration complete! {success_count} table(s) copied successfully.", 'SUCCESS')
        else:
            log(f"Migration finished with errors. OK: {success_count} / Failed: {error_count}.", 'WARN')
        finish(error_count == 0)

    except Exception as exc:
        with _migration_lock:
            _migration_state['logs'].append(f'[ERROR] Unexpected error: {exc}')
        finish(False)


# ─────────────────────────────────────────────────────────────
#  Reproduce DB — Routes
# ─────────────────────────────────────────────────────────────

@app.route('/reproduce-db')
def reproduce_db_page():
    return render_template('reproduce_db.html')


@app.route('/reproduce/tables', methods=['POST'])
def reproduce_tables():
    """Fetch all BASE TABLE names from the source DB (used by Step 1 on the frontend)."""
    data = request.get_json() or {}
    src_conn_str = (data.get('src_conn_str') or '').strip()
    src_db_name  = (data.get('src_db_name')  or '').strip()
    if not src_conn_str or not src_db_name:
        return jsonify({'status': 'error', 'message': 'src_conn_str và src_db_name là bắt buộc.'})
    try:
        upper = src_conn_str.upper()
        if 'DATABASE=' not in upper and 'INITIAL CATALOG=' not in upper:
            src_full = src_conn_str + f';DATABASE={src_db_name}'
        else:
            src_full = src_conn_str
        conn = pyodbc.connect(src_full, timeout=15)
        df   = pd.read_sql(
            "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_TYPE = 'BASE TABLE' ORDER BY TABLE_NAME",
            conn
        )
        conn.close()
        return jsonify({'status': 'success', 'tables': df['TABLE_NAME'].tolist()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/reproduce/selected-tables', methods=['GET', 'POST'])
def reproduce_selected_tables():
    """Persist the user's checked table list so subsequent fetches can restore it."""
    path = os.path.join(get_script_dir(), _SELECTED_TABLES_FILE)
    if request.method == 'GET':
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return jsonify({'status': 'success', 'selected': data.get('selected', [])})
        except FileNotFoundError:
            return jsonify({'status': 'success', 'selected': []})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e), 'selected': []})
    body     = request.get_json() or {}
    selected = body.get('selected', [])
    if not isinstance(selected, list):
        return jsonify({'status': 'error', 'message': 'selected must be a list.'})
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump({'selected': selected}, f, ensure_ascii=False, indent=2)
        return jsonify({'status': 'success', 'message': f'{len(selected)} bảng đã được lưu.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/reproduce/config', methods=['GET', 'POST'])
def reproduce_config():
    if request.method == 'GET':
        return jsonify({'status': 'success', 'config': _get_reproduce_config()})

    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'message': 'No data provided.'})
    try:
        _save_reproduce_config(data)
        return jsonify({'status': 'success', 'message': 'Configuration saved successfully.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Failed to save configuration: {e}'})


@app.route('/reproduce/start', methods=['POST'])
def reproduce_start():
    global _migration_state

    with _migration_lock:
        if _migration_state.get('running'):
            return jsonify({'status': 'error', 'message': 'A migration is already in progress.'})
        _migration_state = {
            'running':  True,
            'done':     False,
            'success':  False,
            'progress': 0,
            'label':    'Starting…',
            'logs':     [],
        }

    # Merge posted values on top of saved config so UI changes take effect immediately
    posted = request.get_json() or {}
    cfg = _get_reproduce_config()
    from datetime import datetime
    for k in _REPRODUCE_CONFIG_DEFAULTS:
        if posted.get(k) not in (None, ''):
            cfg[k] = posted[k]
    # Auto-generate dst_db_name if not provided
    if not cfg.get('dst_db_name'):
        now = datetime.now().strftime('%Y%m%d%H%M')
        cfg['dst_db_name'] = f"db_{now}"
    # Pass selected table list directly (not stored in _REPRODUCE_CONFIG_DEFAULTS)
    if posted.get('selected_tables'):
        cfg['selected_tables'] = posted['selected_tables']

    threading.Thread(target=_run_migration, args=(cfg,), daemon=True).start()
    return jsonify({'status': 'success', 'message': f"Migration started. New Local DB: {cfg['dst_db_name']}"})


@app.route('/reproduce/status', methods=['GET'])
def reproduce_status():
    with _migration_lock:
        done     = _migration_state['done']
        success  = _migration_state['success']
        progress = _migration_state['progress']
        label    = _migration_state['label']
        logs     = list(_migration_state['logs'])
        _migration_state['logs'] = []   # clear: each poll returns only new lines
    return jsonify({
        'done':     done,
        'success':  success,
        'progress': progress,
        'label':    label,
        'logs':     logs,
    })


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5021)
