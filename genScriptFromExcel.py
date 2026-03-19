import os
import re
import pandas as pd
import json
import datetime
from openpyxl import load_workbook
import datetime
import json
import re
import pandas as pd

def join_sql_values(values):
    """
    Join SQL values, handling both tuples and strings.
    Also checks for consecutive commas (,,) and replaces with ,'',
    """
    result_values = []
    for val in values:
        if isinstance(val, tuple):
            # If it's a tuple, take the first element (the value)
            result_values.append(str(val[0]) if val[0] is not None else "''")
        else:
            # If it's a string, use it directly
            result_values.append(str(val) if val is not None else "''")
    
    # Join values with commas
    joined = ", ".join(result_values)
    
    # Check for consecutive commas and replace with ,'',
    while ",," in joined:
        joined = joined.replace(",,", ",'',")
    
    return joined


# Global variables for system id, date, and SEQ per sheet
# Lấy systemid_value từ input, nếu không nhập thì lấy mặc định
now = datetime.datetime.now()
default_systemid_value = f"{now.hour:02d}{now.minute:02d}{now.second:02d}"
systemid_value = default_systemid_value
system_date_value = now.strftime('%Y-%m-%d')

# seq_per_sheet_dict: {sheet_index: SEQ}
seq_per_sheet_dict = {}

# Global workbook and sheetnames - will be initialized once
wb = None
sheetnames = None

# Global table_info - will be initialized once
table_info = None

# Performance optimization caches
_merged_cell_cache = {}  # Cache for merged cell checks: {(sheet_name, row, col_start, col_end): bool}
_cell_value_cache = {}  # Cache for cell values: {(sheet_name, cell_ref): value}
_regex_pattern_cache = {}  # Cache for compiled regex patterns
_sheet_b2_values_cache = {}  # Cache for B2 values: {sheet_name: b2_value}
_username_id_counter = None  # Cache for username ID counter

def clear_performance_caches():
    """Clear all performance caches to free memory"""
    global _merged_cell_cache, _cell_value_cache, _regex_pattern_cache, _sheet_b2_values_cache, _username_id_counter
    _merged_cell_cache.clear()
    _cell_value_cache.clear()
    _regex_pattern_cache.clear()
    _sheet_b2_values_cache.clear()
    _username_id_counter = None

def create_insert_statement_batch(table_name, columns, values_list):
    """Create INSERT statements in batch for better performance"""
    if not values_list:
        return []
    
    columns_str = ", ".join(columns)
    insert_statements = []
    
    for values in values_list:
        values_str = ", ".join(str(v) for v in values)
        sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
        insert_statements.append(sql)
    
    return insert_statements

def preload_sheet_cell_values(ws, start_row, end_row, columns):
    """Pre-load cell values for a range to improve performance"""
    for row in range(start_row, min(end_row + 1, ws.max_row + 1)):
        # Pre-load B column values (commonly used)
        cache_key = (ws.title, f"B{row}")
        if cache_key not in _cell_value_cache:
            try:
                _cell_value_cache[cache_key] = ws[f"B{row}"].value
            except:
                _cell_value_cache[cache_key] = None
        
        # Pre-load other commonly used cell references
        for col in ['C', 'D', 'E']:
            cache_key = (ws.title, f"{col}{row}")
            if cache_key not in _cell_value_cache:
                try:
                    _cell_value_cache[cache_key] = get_cell_value_with_merged(ws, f"{col}{row}")
                except:
                    _cell_value_cache[cache_key] = None

# Script directory for absolute file paths
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Default values for all configurable constants
_DEFAULT_MAPPING_VALUE_DICT = {
    '項目定義書_帳票': '2',
    '項目定義書_画面': '1',
    '項目定義書_CSV': '5',
    '項目定義書_IPO図': '4',
    '項目定義書_ﾒﾆｭｰ': '3',
}
_DEFAULT_KOUMOKU_TYPE_MAPPING = {
    'ラベル': '101',
    'タイトルラベル': '102',
    'テキストボックス': '103',
    'コンボボックス': '104',
    'ラジオボタン': '105',
    'チェックボックス': '106',
    'チェックリスト': '107',
    'ボタン': '108',
    'ボタングループ': '109',
    'メニュートゥール': '110',
    'メニューツリー': '111',
    '画像': '112'
}
_DEFAULT_KOUMOKU_TYPE_MAPPING_RE = {
    'ラベル': '101',
    'タイトルラベル': '102',
    'テキストボックス': '103',
    'チェックボックス': '106',
    'データグリッド': '107',
    '処理': '114',
    'レイアウト': '115',
    '画像': '116'
}
_DEFAULT_STOP_VALUES = [
    '【帳票データ】',
    '【ファンクション定義】',
    '【メッセージ定義】',
    '【タブインデックス定義】',
    '【CSVデータ】',
    '【備考】',
    '【運用上の注意点】',
    '【項目定義】',
    '【一覧定義】',
    '【表示位置定義】'
]
_DEFAULT_EXCLUDED_SHEETNAMES = [
    'カスタマイズ設計書(鑑)', 'カスタマイズ設計書', 'はじめに', '変更履歴'
]

# Configurable constants — loaded from genscript_config.json at startup
MAPPING_VALUE_DICT      = dict(_DEFAULT_MAPPING_VALUE_DICT)
KOUMOKU_TYPE_MAPPING    = dict(_DEFAULT_KOUMOKU_TYPE_MAPPING)
KOUMOKU_TYPE_MAPPING_RE = dict(_DEFAULT_KOUMOKU_TYPE_MAPPING_RE)
STOP_VALUES             = set(_DEFAULT_STOP_VALUES)
EXCLUDED_SHEETNAMES     = set(_DEFAULT_EXCLUDED_SHEETNAMES)


def reload_config():
    """Reload configurable constants from genscript_config.json.
    Called at module load and after the user saves config via the web UI.
    """
    global MAPPING_VALUE_DICT, KOUMOKU_TYPE_MAPPING, KOUMOKU_TYPE_MAPPING_RE
    global STOP_VALUES, EXCLUDED_SHEETNAMES
    try:
        config_path = os.path.join(_SCRIPT_DIR, 'genscript_config.json')
        with open(config_path, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        MAPPING_VALUE_DICT      = cfg.get('mapping_value_dict',      _DEFAULT_MAPPING_VALUE_DICT)
        KOUMOKU_TYPE_MAPPING    = cfg.get('koumoku_type_mapping',    _DEFAULT_KOUMOKU_TYPE_MAPPING)
        KOUMOKU_TYPE_MAPPING_RE = cfg.get('koumoku_type_mapping_re', _DEFAULT_KOUMOKU_TYPE_MAPPING_RE)
        STOP_VALUES             = set(cfg.get('stop_values',         _DEFAULT_STOP_VALUES))
        EXCLUDED_SHEETNAMES     = set(cfg.get('excluded_sheetnames', _DEFAULT_EXCLUDED_SHEETNAMES))
    except Exception:
        pass  # Keep defaults if file is missing or invalid


# Load from file on module init
reload_config()

# Constants for column ranges in merged cell checks
MERGED_CELL_RANGES = {
    'B_TO_BN': (2, 66),
    'B_TO_C': (2, 3),
    'E_TO_AZ': (5, 52),
    'B_TO_D': (2, 4),
    'B_TO_K': (2, 11),
    'E_TO_BN': (5, 66),
    'E_TO_BK': (5, 63),
    'D_TO_O': (4, 15),
    'D_TO_N': (4, 14)
}

# Constants for specific cell values that should be skipped
SKIP_CELL_VALUES = {
    'SCREEN_NUMBER': ['画面', '番号'],
    'MESSAGE_CODE': ['ﾒｯｾｰｼﾞ', 'ｺｰﾄﾞ'],
    'DEFINITION_LOCATION': ['定義場所'],
    'DEFINITION_CATEGORY': ['定義区分'],
    'IPO_HEADER': ['入力画面']
}

# Configuration for different row processors
ROW_PROCESSOR_CONFIG = {
    'koumoku': {
        'table_name': 'T_KIHON_PJ_KOUMOKU',
        'logic_table_name': 'T_KIHON_PJ_KOUMOKU_LOGIC',
        'cell_b_value': '【項目定義】',
        'column_value_processor': 'koumoku_set_value',
        'logic_processor': 'koumoku_logic',
        'seq_prefix': 'SEQ_K'
    },
    'func': {
        'table_name': 'T_KIHON_PJ_FUNC',
        'logic_table_name': 'T_KIHON_PJ_FUNC_LOGIC',
        'cell_b_value': '【ファンクション定義】',
        'column_value_processor': 'func_set_value',
        'logic_processor': 'func_logic',
        'seq_prefix': 'SEQ_F'
    },
    're': {
        'table_name': 'T_KIHON_PJ_KOUMOKU_RE',
        'logic_table_name': 'T_KIHON_PJ_KOUMOKU_RE_LOGIC',
        'cell_b_value': '【項目定義】',
        'column_value_processor': 're_set_value',
        'logic_processor': 're_logic',
        'seq_prefix': 'SEQ_RE'
    },
    'csv': {
        'table_name': 'T_KIHON_PJ_KOUMOKU_CSV',
        'logic_table_name': 'T_KIHON_PJ_KOUMOKU_CSV_LOGIC',
        'cell_b_value': '【項目定義】',
        'column_value_processor': 'csv_set_value',
        'logic_processor': 'csv_logic',
        'seq_prefix': 'SEQ_CSV'
    },
    'message': {
        'table_name': 'T_KIHON_PJ_MESSAGE',
        'cell_b_value': '【メッセージ定義】',
        'column_value_processor': 'message_set_value',
        'seq_prefix': 'SEQ_MS'
    },
    'youken': {
      'table_name': 'T_KIHON_PJ_GAMEN_YOUKEN',
      'column_value_processor': 'youken_set_value',
      'seq_prefix': 'SEQ_Y'
    },
    'tab': {
        'table_name': 'T_KIHON_PJ_TAB',
        'cell_b_value': '【タブインデックス定義】',
        'column_value_processor': 'tab_set_value',
        'seq_prefix': 'SEQ_T'
    },
    'hyouji': {
        'table_name': 'T_KIHON_PJ_HYOUJI',
        'cell_b_value': '【表示位置定義】',
        'column_value_processor': 'hyouji_set_value',
        'seq_prefix': 'SEQ_H'
    },
    'ichiran': {
        'table_name': 'T_KIHON_PJ_ICHIRAN',
        'cell_b_value': '【一覧定義】',
        'column_value_processor': 'ichiran_set_value',
        'seq_prefix': 'SEQ_I'
    },
    'menu': {
        'table_name': 'T_KIHON_PJ_MENU',
        'cell_b_value': '【メニュー定義】',
        'column_value_processor': 'menu_set_value',
        'seq_prefix': 'SEQ_M'
    },
    'ipo': {
        'table_name': 'T_KIHON_PJ_IPO',
        'cell_b_value': '入力画面',
        'column_value_processor': 'ipo_set_value',
        'seq_prefix': 'SEQ_IPO'
    }
}

# Configuration for logic processors
LOGIC_PROCESSOR_CONFIG = {
    'koumoku_logic': {
        'table_name': 'T_KIHON_PJ_KOUMOKU_LOGIC',
        'column_value_processor': 'koumoku_set_value',
        'seq_counter_name': 'SEQ_K_L',
        'cell_b_value': '【項目定義】'
    },
    'func_logic': {
        'table_name': 'T_KIHON_PJ_FUNC_LOGIC',
        'column_value_processor': 'func_set_value',
        'seq_counter_name': 'SEQ_F_L',
        'cell_b_value': '【ファンクション定義】'
    },
    're_logic': {
        'table_name': 'T_KIHON_PJ_KOUMOKU_RE_LOGIC',
        'column_value_processor': 're_set_value',
        'seq_counter_name': 'SEQ_RE_L',
        'cell_b_value': '【項目定義】'
    },
    'csv_logic': {
        'table_name': 'T_KIHON_PJ_KOUMOKU_CSV_LOGIC',
        'column_value_processor': 'csv_set_value',
        'seq_counter_name': 'SEQ_CSV_L',
        'cell_b_value': '【項目定義】'
    }
}

def initialize_workbook(excel_file):
    """
    Initialize global workbook and sheetnames from Excel file.
    Should be called once at the beginning of processing.
    """
    global wb, sheetnames, _merged_cell_cache, _cell_value_cache, _sheet_b2_values_cache
    wb = load_workbook(excel_file, data_only=True)
    sheetnames = wb.sheetnames
    
    # Clear caches
    _merged_cell_cache.clear()
    _cell_value_cache.clear()
    _sheet_b2_values_cache.clear()
    
    # Pre-cache B2 values for all sheets
    for sheet_name in sheetnames:
        if sheet_name not in EXCLUDED_SHEETNAMES:
            try:
                ws = wb[sheet_name]
                b2_value = ws["B2"].value
                _sheet_b2_values_cache[sheet_name] = b2_value
            except Exception:
                _sheet_b2_values_cache[sheet_name] = None
    
    print(f"Initialized workbook with {len(sheetnames)} sheets and cached B2 values")


def initialize_table_info(table_info_file):
    """
    Initialize global table_info from JSON file.
    Should be called once at the beginning of processing.
    """
    global table_info
    table_info = read_table_info(table_info_file)
    print(f"Initialized table_info with {len(table_info)} tables")


def read_table_info(filename):
    """
    Reads the JSON content from the given filename and returns it as a dictionary.
    """
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def get_cell_value_with_merged(ws, cell_ref):
    """Helper function to get cell value considering merged cells with caching"""
    cache_key = (ws.title, cell_ref)
    
    # Check cache first
    if cache_key in _cell_value_cache:
        return _cell_value_cache[cache_key]
    
    cell = ws[cell_ref]
    if cell.value is not None:
        _cell_value_cache[cache_key] = cell.value
        return cell.value
    
    # If cell is empty, check merged cells
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            value = ws[merged_range.start_cell.coordinate].value
            _cell_value_cache[cache_key] = value
            return value
    
    _cell_value_cache[cache_key] = None
    return None

def is_merged_from_to(ws, row, col_start, col_end):
    """Check if cells in a row are merged from col_start to col_end with caching"""
    cache_key = (ws.title, row, col_start, col_end)
    
    # Check cache first
    if cache_key in _merged_cell_cache:
        return _merged_cell_cache[cache_key]
    
    result = any(
        merged_range.min_col == col_start
        and merged_range.max_col == col_end
        and merged_range.min_row <= row <= merged_range.max_row
        for merged_range in ws.merged_cells.ranges
    )
    
    _merged_cell_cache[cache_key] = result
    return result


def should_stop_logic_row(ws, check_row, stop_values, cell_b_value=''):
    """Determine action for logic row processing - Simplified and more permissive"""
    if check_row > ws.max_row:
        return 'stop'
    else:
        cell_b_check = ws[f"B{check_row}"].value
        if cell_b_check is None:
            cell_b_check = ''
        merged_b_to_bn = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_BN'])
        merged_bc = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_C'])
        
        if merged_bc:
            # Check for specific skip values
            if cell_b_check in SKIP_CELL_VALUES['SCREEN_NUMBER']:
                return 'skip'
            else:
                return 'stop'
        else:
            if merged_b_to_bn:
                return 'continue'
            else:
                # Check stop conditions first
                if cell_b_check in stop_values:
                    if cell_b_check != cell_b_value:
                        return 'stop'
                    else:
                        return 'skip'
                else:
                    # If not merged and not in stop values, continue processing
                    return 'skip'
   
def _handle_item_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【項目定義】 and 【ファンクション定義】"""
    merged_b_to_bn = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_BN'])
    merged_bc = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_C'])
    
    if merged_bc:
        if cell_b_check in SKIP_CELL_VALUES['SCREEN_NUMBER']:
            return 'skip'
        elif not merged_b_to_bn:
            return 'continue'
    else:
        if merged_b_to_bn:
            # Only create logic on the first logic row encountered
            return 'create_logic'
        else:
            return 'skip'
    return 'skip'

def _handle_message_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【メッセージ定義】"""
    merged_e_to_az = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['E_TO_AZ'])
    merged_bd = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_D'])

    if merged_bd:
        if cell_b_check in SKIP_CELL_VALUES['MESSAGE_CODE']:
            return 'skip'
        return 'continue' if merged_e_to_az else 'skip'
    return 'skip'

def _handle_tab_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【タブインデックス定義】"""
    merged_e_to_bn = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['E_TO_BN'])
    merged_bd = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_D'])

    if merged_bd:
        if cell_b_check in SKIP_CELL_VALUES['DEFINITION_LOCATION']:
            return 'skip'
        return 'continue' if merged_e_to_bn else 'skip'
    return 'skip'

def _handle_position_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【表示位置定義】"""
    merged_e_to_bk = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['E_TO_BK'])
    merged_bd = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_D'])

    if merged_bd:
        if cell_b_check in SKIP_CELL_VALUES['DEFINITION_CATEGORY']:
            return 'skip'
        return 'continue' if merged_e_to_bk else 'skip'
    return 'skip'

def _handle_list_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【一覧定義】"""
    merged_d_to_o = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['D_TO_O'])
    merged_bc = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_C'])

    if merged_bc:
        if cell_b_check in SKIP_CELL_VALUES['SCREEN_NUMBER']:
            return 'skip'
        return 'continue' if merged_d_to_o else 'skip'
    return 'skip'

def _handle_menu_definition_check(ws, check_row, cell_b_check):
    """Handle logic for 【メニュー定義】"""
    merged_d_to_n = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['D_TO_N'])
    merged_bc = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_C'])

    if merged_bc:
        if cell_b_check in SKIP_CELL_VALUES['SCREEN_NUMBER']:
            return 'skip'
        return 'continue' if merged_d_to_n else 'skip'
    return 'skip'

def _handle_ipo_definition_check(ws, check_row, cell_b_check):
    """Handle logic for IPO"""
    merged_b_to_bn = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_BN'])
    merged_b_to_k = is_merged_from_to(ws, check_row, *MERGED_CELL_RANGES['B_TO_K'])
    
    if merged_b_to_k or merged_b_to_bn:
        if '入力画面' in SKIP_CELL_VALUES['SCREEN_NUMBER']:
            return 'skip'
        return 'continue'
    return 'skip'

def should_stop_row(ws, check_row, stop_values, cell_b_value=None):
    """
    Returns action for row processing for T_KIHON_PJ_KOUMOKU_RE (and similar tables):
    1. If cell B is in stop_values (excluding cell_b_value if provided)
    2. End of sheet is handled by the caller
    """
    if check_row > ws.max_row:
        return 'stop'
    
    cell_b_check = ws[f"B{check_row}"].value
    
    # Check stop conditions
    if cell_b_value is not None:
        if cell_b_check in stop_values and cell_b_check != cell_b_value:
            return 'stop'
    else:
        if cell_b_check in stop_values:
            return 'stop'
    
    # Handle specific cell_b_value cases
    if cell_b_value is None:
        return 'skip'
    
    # Simple stop cases
    simple_stop_values = ['【備考】','【運用上の注意点】']
    if cell_b_value in simple_stop_values:
        return 'stop'
    
    # Complex logic cases
    handlers = {
        '【項目定義】': _handle_item_definition_check,
        '【ファンクション定義】': _handle_item_definition_check,
        '【メッセージ定義】': _handle_message_definition_check,
        '【タブインデックス定義】': _handle_tab_definition_check,   
        '【表示位置定義】': _handle_position_definition_check,
        '【一覧定義】': _handle_list_definition_check,
        '【メニュー定義】': _handle_menu_definition_check,
        '入力画面': _handle_ipo_definition_check
    }
    
    handler = handlers.get(cell_b_value)
    if handler:
        return handler(ws, check_row, cell_b_check)
    
    return 'skip'

def set_value_generic(
    col_info,
    ws,
    row_num,
    sheet_seq,
    primary_seq_value,
    secondary_seq_value=None,
    seq_mappings=None,
    reference_mappings=None,
    table_name=None
):
    """
    Refactored generic function to process column values for all table types.
    This version improves readability and manageability by modularizing logic.
    Returns tuple (value, aoji) where aoji indicates font color status.
    """
    val_rule = col_info.get('VALUE', '')
    cell_fix = col_info.get('CELL_FIX', '').strip()
    col_logic = col_info.get('CELL_LOGIC', '').strip()
    col_name = col_info.get('COLUMN_NAME', '')
    data_type = col_info.get('DATA_TYPE', '').lower()
    aoji = False

    def get_seq_value():
        if seq_mappings and col_name in seq_mappings:
            seq_val = seq_mappings[col_name]
            return str(seq_val) if seq_val is not None else "''"
        return None

    def get_seq_reference_value():
        if reference_mappings and val_rule in reference_mappings:
            ref_val = reference_mappings[val_rule]
            return str(ref_val) if ref_val is not None else "''"
        return None

    def handle_mapping():
        nonlocal aoji
        mapped_val = ''
        if cell_fix:
            cell_ref = f"{cell_fix}{row_num}"
            cell_value = ws[cell_ref].value if ws[cell_ref].value else None
            # Extract font color for aoji
            try:
                cell = ws[cell_ref]
                font_rgb = get_font_rgb(cell)
                aoji = is_aoji(font_rgb)
            except Exception:
                pass
        else:
            cell_ref = f"{col_logic}{row_num}"
            cell_value = get_cell_value_with_merged(ws, cell_ref)
            # Extract font color for aoji
            try:
                cell = ws[cell_ref]
                font_rgb = get_font_rgb(cell)
                aoji = is_aoji(font_rgb)
            except Exception:
                pass
            if col_name == 'KOUMOKU_SYURUI_CD' and isinstance(cell_value, str):
                if table_name == 'T_KIHON_PJ_KOUMOKU':
                    mapped_val = KOUMOKU_TYPE_MAPPING.get(cell_value, '')
                elif table_name == 'T_KIHON_PJ_KOUMOKU_RE':
                    mapped_val = KOUMOKU_TYPE_MAPPING_RE.get(cell_value, '')

        return f"'{mapped_val}'" if mapped_val else "''"

    def get_font_rgb(cell):
        # openpyxl >= 2.5: cell.font.color is a Color object, .rgb is a string or None
        if cell.font and cell.font.color:
            rgb = getattr(cell.font.color, 'rgb', None)
            # Only accept if rgb is a string of length 6 or 8 (hex color)
            if isinstance(rgb, str) and (len(rgb) == 6 or len(rgb) == 8):
                return rgb.upper()
        return None

    def is_aoji(font_rgb):
        # Nếu font_rgb là None (mặc định, không tô màu), coi là đen (aoji=False)
        black_colors = {None, '000000', 'FF000000'}
        if font_rgb is None:
            return False
        return font_rgb not in black_colors

    # Handle AUTO_ID cases with sequence mappings
    if val_rule == 'AUTO_ID':
        seq_val = get_seq_value()
        if seq_val is not None:
            return seq_val, aoji

    # Handle specific reference mappings
    ref_val = get_seq_reference_value()
    if ref_val is not None:
        return ref_val, aoji

    # Handle MAPPING case
    if val_rule == 'MAPPING':
        return handle_mapping(), aoji

    # Handle empty value rule (direct cell reading)
    if val_rule == '':
        try:
            if cell_fix:
                cell_ref = f"{cell_fix}{row_num}"
                cell = ws[cell_ref]
                cell_value = get_cell_value_with_merged(ws, cell_ref)
                font_rgb = get_font_rgb(cell)
                aoji = is_aoji(font_rgb)
            elif col_logic:
                cell_ref = f"{col_logic}{row_num}"
                cell = ws[cell_ref]
                cell_value = get_cell_value_with_merged(ws, cell_ref)
                font_rgb = get_font_rgb(cell)
                aoji = is_aoji(font_rgb)
                # Special case for YOUKEN_NO pattern extraction
                if col_name == 'YOUKEN_NO':
                    extracted_value = _extract_youken_no(cell_value)
                    if extracted_value:
                        cell_value = extracted_value
                    else:
                        return "''", aoji
                # Special case for MIDASHI: if B~BN is merged at this row, set cell_value = 'True'
                if col_name == 'MIDASHI':
                    if is_merged_from_to(ws, row_num, 2, 66):  # B=2, BN=66
                        cell_value = 'True'
                    else:
                        cell_value = 'False'
            else:
                cell_value = None
            return _format_cell_value_by_type(cell_value, data_type, col_name, table_name), aoji
        except Exception:
            return "''", aoji

    # Handle T_KIHON_PJ_GAMEN.SEQ reference
    if val_rule == 'T_KIHON_PJ_GAMEN.SEQ':
        return str(sheet_seq) if sheet_seq is not None else "''", aoji

    val = "''"
    if val_rule == 'BLANK':
        val = "''"
    elif val_rule == 'NULL':
        val = "NULL"
    elif val_rule == 'SYSTEMID':
        val = f"'{systemid_value}'"
    elif val_rule == 'T_KIHON_PJ.SYSTEM_ID':
        val = f"'{systemid_value}'"
    return val if val else "''", aoji

def _format_cell_value_by_type(cell_value, data_type, col_name=None, table_name=None):
    """Format cell value based on data type"""
    if cell_value is None or cell_value == '':
        return "''"
    
    # Additional handling for specific columns in T_KIHON_PJ_KOUMOKU
    if (
        col_name in ['ZENKAKU_MOJI_SU', 'HANKAKU_MOJI_SU', 'SEISU_KETA', 'SYOUSU_KETA'] and
        cell_value == "－"
    ):
        return "NULL"
   
    # Numeric types: do not quote, return as int/float
    if data_type in ['int'] and cell_value != "NULL":
        return int(cell_value)

    # Date/time types: quote and format
    elif data_type in ['date', 'datetime', 'smalldatetime', 'datetime2', 'datetimeoffset', 'time']:
        if isinstance(cell_value, datetime.datetime):
            return f"'{cell_value.strftime('%Y-%m-%d %H:%M:%S')}'"
        elif isinstance(cell_value, str):
            return f"'{cell_value}'"
        else:
            return f"'{str(cell_value)}'"
    # NVARCHAR: N'...'
    elif data_type == 'nvarchar':
        return f"N'{cell_value}'"
    # Default: quote as string
    else:
        return f"'{cell_value}'"

def _extract_youken_no(cell_value):
    """Extract YOUKEN_NO pattern from cell value with caching"""
    if isinstance(cell_value, str):
        # Use cached regex pattern
        pattern_key = 'youken_pattern'
        if pattern_key not in _regex_pattern_cache:
            _regex_pattern_cache[pattern_key] = re.compile(r'^\(要件№([\d\-]+)\)要件ﾛｼﾞｯｸ：')
        
        pattern = _regex_pattern_cache[pattern_key]
        m = pattern.match(cell_value)
        if m:
            return m.group(1)
    return None




def koumoku_set_value(col_info, ws, row_num, sheet_seq, seq_k_value, seq_k_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU table"""
    seq_mappings = {
        'SEQ_K': seq_k_value,
        'ROW_NO': seq_k_value,
        'SEQ_K_L': seq_k_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU.SEQ_K': seq_k_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_k_value,
        secondary_seq_value=seq_k_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        table_name='T_KIHON_PJ_KOUMOKU'
    )


def func_set_value(col_info, ws, row_num, sheet_seq, seq_f_value, seq_f_l_value=None):
    """Process column value for T_KIHON_PJ_FUNC table"""
    seq_mappings = {
        'SEQ_F': seq_f_value,
        'ROW_NO': seq_f_value,
        'SEQ_F_L': seq_f_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_FUNC.SEQ_F': seq_f_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_f_value,
        secondary_seq_value=seq_f_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings
    )


def csv_set_value(col_info, ws, row_num, sheet_seq, seq_csv_value, seq_csv_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU_CSV table"""
    seq_mappings = {
        'SEQ_CSV': seq_csv_value,
        'ROW_NO': seq_csv_value,
        'SEQ_CSV_L': seq_csv_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU_CSV.SEQ_CSV': seq_csv_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_csv_value,
        secondary_seq_value=seq_csv_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings
    )


def re_set_value(col_info, ws, row_num, sheet_seq, seq_re_value, seq_re_l_value=None):
    """Process column value for T_KIHON_PJ_KOUMOKU_RE table"""
    seq_mappings = {
        'SEQ_RE': seq_re_value,
        'ROW_NO': seq_re_value,
        'SEQ_RE_L': seq_re_l_value
    }
    reference_mappings = {
        'T_KIHON_PJ_KOUMOKU_RE.SEQ_RE': seq_re_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_re_value,
        secondary_seq_value=seq_re_l_value,
        seq_mappings=seq_mappings,
        reference_mappings=reference_mappings,
        table_name='T_KIHON_PJ_KOUMOKU_RE'
    )


def message_set_value(col_info, ws, row_num, sheet_seq, seq_ms_value):
    """Process column value for T_KIHON_PJ_MESSAGE table"""
    seq_mappings = {
        'SEQ_MS': seq_ms_value,
        'ROW_NO': seq_ms_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_ms_value,
        seq_mappings=seq_mappings
    )

def youken_set_value(col_info, ws, row_num, sheet_seq, seq_yk_value):
    # Không có ROW_NO, không có AOJI
    seq_mappings = {
        'SEQ_Y': seq_yk_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_yk_value,
        seq_mappings=seq_mappings
    )

def insert_youken_from_S7(ws, sheet_seq):
    col_info_list = table_info['T_KIHON_PJ_GAMEN_YOUKEN']
    s7_value = ws['S7'].value or ''
    lines = [line.replace('_x000D_', '').strip() for line in str(s7_value).splitlines() if line.strip()]
    insert_statements = []
    for idx, line in enumerate(lines, 1):
        # Tách line thành YOUKEN_NO và YOUKEN_GAIYOU
        if ':' in line:
            youken_no, youken_gaiyou = line.split(':', 1)
        else:
            youken_no, youken_gaiyou = '', line  # Nếu không có dấu :, coi toàn bộ là GAIYOU

        row_data = {}
        for col_info in col_info_list:
            col_name = col_info['COLUMN_NAME']
            if col_name == 'YOUKEN_NO':
                val = f"N'{youken_no.strip()}'"
                aoji = False
            elif col_name == 'YOUKEN_GAIYOU':
                val = f"N'{youken_gaiyou.strip()}'"
                aoji = False
            else:
                val, aoji = youken_set_value(col_info, ws, 7, sheet_seq, idx)
            row_data[col_name] = val
        columns_str = ", ".join(row_data.keys())
        values_str = join_sql_values(row_data.values())
        sql = f"INSERT INTO T_KIHON_PJ_GAMEN_YOUKEN ({columns_str}) VALUES ({values_str});"
        insert_statements.append(sql)
    return insert_statements

def hyouji_set_value(col_info, ws, row_num, sheet_seq, seq_hj_value):
    """Process column value for T_KIHON_PJ_HYOUJI table"""
    seq_mappings = {
        'SEQ_H': seq_hj_value,
        'ROW_NO': seq_hj_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_hj_value,
        seq_mappings=seq_mappings
    )
    
def tab_set_value(col_info, ws, row_num, sheet_seq, seq_t_value):
    """Process column value for T_KIHON_PJ_TAB table"""
    seq_mappings = {
        'SEQ_T': seq_t_value,
        'ROW_NO': seq_t_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_t_value,
        seq_mappings=seq_mappings
    )


def ichiran_set_value(col_info, ws, row_num, sheet_seq, seq_i_value):
    """Process column value for T_KIHON_PJ_ICHIRAN table"""
    seq_mappings = {
        'SEQ_I': seq_i_value,
        'ROW_NO': seq_i_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_i_value,
        seq_mappings=seq_mappings
    )


def menu_set_value(col_info, ws, row_num, sheet_seq, seq_m_value):
    """Process column value for T_KIHON_PJ_MENU table"""
    seq_mappings = {
        'SEQ_M': seq_m_value,
        'ROW_NO': seq_m_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_m_value,
        seq_mappings=seq_mappings
    )


def ipo_set_value(col_info, ws, row_num, sheet_seq, seq_ipo_value):
    """Process column value for T_KIHON_PJ_IPO table"""
    seq_mappings = {
        'SEQ_IPO': seq_ipo_value,
        'ROW_NO': seq_ipo_value
    }
    
    return set_value_generic(
        col_info=col_info,
        ws=ws,
        row_num=row_num,
        sheet_seq=sheet_seq,
        primary_seq_value=seq_ipo_value,
        seq_mappings=seq_mappings
    )


def _handle_username_id(cell_value):
    """Handle USER_NAME ID generation from usernameID.txt file with caching"""
    global _username_id_counter
    
    # Initialize counter if not cached
    _username_id_path = os.path.join(_SCRIPT_DIR, 'usernameID.txt')
    if _username_id_counter is None:
        try:
            with open(_username_id_path, 'r', encoding='utf-8') as f:
                current_id = f.read().strip()
                if not current_id.isdigit():
                    current_id = '1'
                _username_id_counter = int(current_id)
        except Exception:
            _username_id_counter = 1
    
    # Use current counter, then decrease by 1
    current_id = _username_id_counter
    _username_id_counter -= 1
    if _username_id_counter < 1:
        _username_id_counter = 1
    
    # Write new value to file only when changed
    try:
        new_id = str(_username_id_counter).zfill(len(str(current_id)))
        with open(_username_id_path, 'w', encoding='utf-8') as f:
            f.write(new_id)
    except Exception:
        pass
    
    return str(cell_value) + str(current_id)

def _parse_ref_pattern(ref_value):
    """Parse REF pattern like 'G2' into column letter and row number with caching"""
    if not ref_value or not isinstance(ref_value, str):
        return None, None
    
    # Use cached regex pattern
    pattern_key = 'ref_pattern'
    if pattern_key not in _regex_pattern_cache:
        _regex_pattern_cache[pattern_key] = re.compile(r'^([A-Z]+)(\d+)$')
    
    pattern = _regex_pattern_cache[pattern_key]
    match = pattern.match(ref_value.strip().upper())
    if match:
        return match.group(1), int(match.group(2))
    return None, None

def _find_ref_data_row(ws, target_value, stop_values=None):
    """Find row where column B contains target_value, stop at stop_values"""
    if stop_values is None:
        stop_values = ['【備考】', '【運用上の注意点】']
    
    for row_num in range(1, ws.max_row + 1):
        cell_b_value = ws[f"B{row_num}"].value
        if cell_b_value == target_value:
            return row_num
        if cell_b_value in stop_values:
            break
    return None

def _get_ref_cell_value(ws, sheet_check_value, ref_value, col_name):
    """Get cell value based on REF pattern and sheet type"""
    # Parse REF pattern
    col_letter, row_offset = _parse_ref_pattern(ref_value)
    if not col_letter or row_offset is None:
        return 'NULL'
    
    # Determine target value to search for based on sheet type
    target_mapping = {
        '項目定義書_画面': '【抽出データ定義】',
        '項目定義書_帳票': '【帳票データ】', 
        '項目定義書_CSV': '【CSVデータ】'
    }
    
    target_value = target_mapping.get(sheet_check_value)
    if not target_value:
        return 'NULL'
    
    # Find the target row
    target_row = _find_ref_data_row(ws, target_value)
    if not target_row:
        return 'NULL'
    
    # Calculate final cell position: [X][target_row + 1 + Y*2]
    final_row = target_row + 1 + (row_offset * 2)
    cell_ref = f"{col_letter}{final_row}"
    
    try:
        cell_value = get_cell_value_with_merged(ws, cell_ref)
        if col_name == 'KUGIRI_MOJI_KB_CSV' and cell_value is not None:
            if cell_value == 'カンマ':
                return 'カンマ'
            elif cell_value == 'タブ':
                return 'タブ'
            else:
                return'その他'
            
        return cell_value if cell_value is not None else 'NULL'
    except Exception:
        return 'NULL'

def _format_value_by_data_type(cell_value, data_type, col_name):
    """Format value based on data type"""
    # Numeric types: do not quote
    if data_type in ['int', 'bigint', 'smallint', 'tinyint', 'decimal', 'numeric', 'float', 'real', 'money', 'smallmoney']:
        try:
            if isinstance(cell_value, (int, float)):
                return str(cell_value)
            else:
                cell_str = str(cell_value).replace(',', '').replace(' ', '')
                if '.' in cell_str:
                    return str(float(cell_str))
                else:
                    return str(int(cell_str))
        except Exception:
            return '0'
    
    # Date/time types: quote and format
    elif data_type in ['date', 'datetime', 'smalldatetime', 'datetime2', 'datetimeoffset', 'time']:
        if isinstance(cell_value, datetime.datetime):
            return f"'{cell_value.strftime('%Y-%m-%d %H:%M:%S')}'"
        elif isinstance(cell_value, str):
            return f"'{cell_value}'"
        else:
            return f"'{str(cell_value)}'"
    
    # NVARCHAR: N'...'
    elif data_type == 'nvarchar':
        if col_name == 'USER_NAME':
            username_with_id = _handle_username_id(cell_value)
            return f"N'{username_with_id}'"
        else:
            return f"N'{cell_value}'"
    
    # Default: quote as string
    else:
        return f"'{cell_value}'"

def column_value(col_info, ws, systemid_value, system_date_value, seq_value=None, jyun_value=None, sheet_check_value=None):
    """Process column value based on VALUE rules"""
    val_rule = col_info.get('VALUE', '')
    cell_fix = col_info.get('CELL_FIX', '').strip()
    col_name = col_info.get('COLUMN_NAME', '')
    aoji = False  # Default font color status
    
    # Extract font color if cell_fix is available
    if cell_fix:
        try:
            cell = ws[cell_fix]
            if cell.font and cell.font.color:
                rgb = getattr(cell.font.color, 'rgb', None)
                if isinstance(rgb, str) and (len(rgb) == 6 or len(rgb) == 8):
                    black_colors = {None, '000000', 'FF000000'}
                    aoji = rgb.upper() not in black_colors
        except Exception:
            pass
    
    if val_rule == 'BLANK':
        val = "''"
    elif val_rule == 'NULL':
        val = "NULL"
    elif val_rule == 'SYSTEMID':
        val = f"'{systemid_value}'"
    elif val_rule == 'T_KIHON_PJ.SYSTEM_ID':
        val = f"'{systemid_value}'"
    elif val_rule == 'AUTO_ID' and col_name == 'SEQ':
        val = str(seq_value) if seq_value is not None else "''"
    elif val_rule == 'AUTO_ID' and col_name == 'JYUN':
        val = str(jyun_value) if jyun_value is not None else "''"
    elif val_rule in ('SYSTEM DATE', 'AUTO_TIME'):
        val = f"'{system_date_value}'"
    elif val_rule == 'MAPPING':
        cell_value = ws[cell_fix].value if cell_fix else None
        val = MAPPING_VALUE_DICT.get(cell_value, "''")
    elif val_rule == 'REF':
        # Handle REF case based on sheet_check_value
        if sheet_check_value in ['項目定義書_画面', '項目定義書_帳票', '項目定義書_CSV']:
            # Determine which column to check based on sheet type
            ref_column_mapping = {
                '項目定義書_画面': 'SCREEN',
                '項目定義書_帳票': 'REPORT', 
                '項目定義書_CSV': 'CSV'
            }
            ref_column = ref_column_mapping.get(sheet_check_value)
            
            if ref_column:
                cell_ref = col_info.get(ref_column, '').strip()
                try:
                    if cell_ref:
                        cell_value = _get_ref_cell_value(ws, sheet_check_value, cell_ref, col_name)
                        if cell_value == 'NULL':
                            val = "NULL"
                        else:
                            data_type = col_info.get('DATA_TYPE', '').lower()
                            val = _format_value_by_data_type(cell_value, data_type, col_name)
                    else:
                        val = "NULL"
                except Exception:
                    val = "NULL"
            else:
                val = "NULL"
        else:
            val = "NULL"
    elif val_rule == '':
        if cell_fix:
            try:
                # Add logic for SHEET_NAME
                if col_name == 'SHEET_NAME':
                    # Try to get sheet_idx from ws
                    global sheetnames
                    sheet_idx = None
                    for idx, name in enumerate(sheetnames):
                        if ws.title == name:
                            sheet_idx = idx
                            break
                    if sheet_idx is not None:
                        cell_value = sheetnames[sheet_idx]
                    else:
                        cell_value = ws.title
                else:
                    cell_value = get_cell_value_with_merged(ws, cell_fix)
                if cell_value is None:
                    val = "''"
                else:
                    data_type = col_info.get('DATA_TYPE', '').lower()
                    val = _format_value_by_data_type(cell_value, data_type, col_name)
            except Exception:
                val = "''"
        else:
            val = "''"
    else:
        # Other values, treat as string literal
        # Add N prefix for nvarchar columns
        if col_info.get('DATA_TYPE', '').lower() == 'nvarchar':
            val = f"N'{val_rule}'"
        else:
            val = f"'{val_rule}'"
    
    return val, aoji
  


def generate_insert_statements_from_excel(sheet_index, table_key):
    """
    Unified function to generate INSERT statements for all table types
    Uses global wb, sheetnames, and table_info instead of loading files each time
    """
    # Use global table_info instead of reading file
    global table_info, wb, sheetnames, systemid_value, system_date_value
    
    if table_key not in table_info:
        raise ValueError(f"Table key '{table_key}' not found in table info.")
    
    columns_info = table_info[table_key]
    insert_statements = []
    
    if table_key == 'T_KIHON_PJ_GAMEN':
        # Special handling for T_KIHON_PJ_GAMEN: process multiple sheets
        global seq_per_sheet_dict
        seq_per_sheet = 1
        allowed_b2_values = set(MAPPING_VALUE_DICT.keys())
        for sheet_idx, sheet_name in enumerate(sheetnames):
            if sheet_name in EXCLUDED_SHEETNAMES:
                continue
            ws = wb[sheet_name]
            try:
                sheet_check_value = ws["B2"].value
            except Exception:
                sheet_check_value = None
            if sheet_check_value not in allowed_b2_values:
                continue
            row_data = {}
            seq_value = seq_per_sheet
            jyun_value = seq_value
            seq_per_sheet_dict[sheet_idx] = seq_value
            aoji_values = []
            for col_info in columns_info:
                col_name = col_info.get('COLUMN_NAME', '')
                val, aoji = column_value(col_info, ws, systemid_value, system_date_value, seq_value, jyun_value)
                row_data[col_name] = val
                aoji_values.append(aoji)
            
            # Set AOJI column based on collected aoji values
            final_aoji = '1' if any(aoji_values) else '0'
            if 'AOJI' in row_data:
                row_data['AOJI'] = f"'{final_aoji}'"
            
            columns_str = ", ".join(row_data.keys())
            values_str = join_sql_values(row_data.values())
            sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
            seq_per_sheet += 1
    
    elif table_key == 'T_KIHON_PJ':
        # Special handling for T_KIHON_PJ: single insert statement
        if sheet_index >= len(sheetnames):
            raise ValueError(f"Sheet index {sheet_index} out of range.")
        ws = wb[sheetnames[sheet_index]]
        
        cols = []
        vals = []
        aoji_values = []
        for col_info in columns_info:
            col_name = col_info['COLUMN_NAME']
            cols.append(col_name)
            val, aoji = column_value(col_info, ws, systemid_value, system_date_value)
            vals.append(val)
            aoji_values.append(aoji)

        # Set AOJI column based on collected aoji values
        final_aoji = '1' if any(aoji_values) else '0'
        if 'AOJI' in cols:
            aoji_index = cols.index('AOJI')
            vals[aoji_index] = f"'{final_aoji}'"

        columns_str = ", ".join(cols)
        values_str = join_sql_values(vals)
        sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
        insert_statements.append(sql)
    
    else:
        # Default handling for other tables: process each row in the sheet
        df = pd.read_excel(wb, sheet_name=sheet_index, engine='openpyxl')
        if sheet_index >= len(sheetnames):
            raise ValueError(f"Sheet index {sheet_index} out of range.")
        ws = wb[sheetnames[sheet_index]]
        
        for _, row in df.iterrows():
            cols = []
            vals = []
            aoji_values = []
            for col_info in columns_info:
                col_name = col_info['COLUMN_NAME']
                cols.append(col_name)
                val, aoji = column_value(col_info, ws, systemid_value, system_date_value)
                vals.append(val)
                aoji_values.append(aoji)
            
            # Set AOJI column based on collected aoji values
            final_aoji = '1' if any(aoji_values) else '0'
            if 'AOJI' in cols:
                aoji_index = cols.index('AOJI')
                vals[aoji_index] = f"'{final_aoji}'"
            
            columns_str = ", ".join(cols)
            values_str = join_sql_values(vals)
            sql = f"INSERT INTO {table_key} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
    
    return insert_statements


def all_tables_in_sequence(excel_file, table_info_file, output_file='insert_all.sql'):
    """
    Process all tables in the correct sequence:
    1. Initialize workbook and table_info once
    2. Create INSERT for T_KIHON_PJ
    3. Iterate through sheets (from sheet 3) to create INSERT for T_KIHON_PJ_GAMEN
    4. For each new SEQ, process T_KIHON_PJ_KOUMOKU
    5. For each new SEQ_K, process T_KIHON_PJ_KOUMOKU_LOGIC
    """
    global seq_per_sheet_dict, wb, sheetnames, table_info
    
    # Initialize workbook and table_info once at the beginning
    initialize_workbook(excel_file)
    initialize_table_info(table_info_file)
    
    all_insert_statements = []
    
    pj_inserted = False
    
    gamen_columns_info = table_info.get('T_KIHON_PJ_GAMEN', [])
    
    seq_per_sheet = 1
    allowed_b2_values = set(MAPPING_VALUE_DICT.keys())
    
    for sheet_idx, sheet_name in enumerate(sheetnames):
        if sheet_name in EXCLUDED_SHEETNAMES:
            continue
        
        # Use cached B2 value instead of reading from sheet
        sheet_check_value = _sheet_b2_values_cache.get(sheet_name)

        if sheet_check_value not in allowed_b2_values:
            continue

        # Lồng logic tạo INSERT cho T_KIHON_PJ, chỉ thực hiện 1 lần cho sheet hợp lệ đầu tiên
        if not pj_inserted:
            print("Processing T_KIHON_PJ...")
            pj_inserts = generate_insert_statements_from_excel(sheet_idx, 'T_KIHON_PJ')
            all_insert_statements.extend(pj_inserts)
            pj_inserted = True

        # Always process T_KIHON_PJ_GAMEN
        ws = wb[sheet_name]  # Get worksheet reference
        row_data = {}
        seq_value = seq_per_sheet
        jyun_value = seq_value
        seq_per_sheet_dict[sheet_idx] = seq_value
        aoji_values = []
        for col_info in gamen_columns_info:
            col_name = col_info.get('COLUMN_NAME', '')
            val, aoji = column_value(col_info, ws, systemid_value, system_date_value, seq_value, jyun_value, sheet_check_value)
            row_data[col_name] = val
            aoji_values.append(aoji)
        
        # Set AOJI column based on collected aoji values
        final_aoji = '1' if any(aoji_values) else '0'
        if 'AOJI' in row_data:
            row_data['AOJI'] = f"'{final_aoji}'"
        
        columns_str = ", ".join(row_data.keys())
        values_str = join_sql_values(row_data.values())
        sql = f"INSERT INTO T_KIHON_PJ_GAMEN ({columns_str}) VALUES ({values_str});"
        all_insert_statements.append(sql)
        print(f"Processing sheet {sheet_idx}: {sheet_name} with SEQ {seq_value}")

        youken_inserts = insert_youken_from_S7(ws, seq_value)
        all_insert_statements.extend(youken_inserts)
        
        
        # Xử lý theo từng loại sheet_check_value
        if sheet_check_value == '項目定義書_帳票':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_KOUMOKU_RE, T_KIHON_PJ_KOUMOKU_RE_LOGIC
            re_inserts = re_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(re_inserts)
        elif sheet_check_value == '項目定義書_CSV':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_KOUMOKU_CSV, T_KIHON_PJ_KOUMOKU_CSV_LOGIC
            csv_inserts = csv_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(csv_inserts)
        elif sheet_check_value == '項目定義書_IPO図':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_IPO
            ipo_inserts = ipo_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(ipo_inserts)
        elif sheet_check_value == '項目定義書_ﾒﾆｭｰ':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_MENU
            menu_inserts = menu_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(menu_inserts)
        elif sheet_check_value == '項目定義書_画面':
            # Chỉ xử lý T_KIHON_PJ_GAMEN, T_KIHON_PJ_FUNC, T_KIHON_PJ_FUNC_LOGIC, T_KIHON_PJ_KOUMOKU, T_KIHON_PJ_KOUMOKU_LOGIC, T_KIHON_PJ_MESSAGE, T_KIHON_PJ_TAB, T_KIHON_PJ_ICHIRAN, T_KIHON_PJ_HYOUJI
            koumoku_inserts = koumoku_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(koumoku_inserts)
            
            func_inserts = func_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(func_inserts)
            
            message_inserts = message_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(message_inserts)
            tab_inserts = tab_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(tab_inserts)
            ichiran_inserts = ichiran_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(ichiran_inserts)
            hyouji_inserts = hyouji_row(
                sheet_idx, seq_value
            )
            all_insert_statements.extend(hyouji_inserts)
        seq_per_sheet += 1
    
    # Write all statements to file in batches for better I/O performance
    batch_size = 1000
    with open(output_file, 'w', encoding='utf-8') as f:
        for i in range(0, len(all_insert_statements), batch_size):
            batch = all_insert_statements[i:i + batch_size]
            f.write('\n'.join(batch) + '\n')
    
    # Clear caches after processing to free memory
    clear_performance_caches()
    
    print(f"All INSERT statements written to {output_file}")
    return all_insert_statements


def gen_row_single_sheet(
    sheet_idx,
    sheet_seq,
    table_name,
    logic_table_name=None,
    cell_b_value='【項目定義】',
    column_value_processor=None,
    logic_processor=None,
    seq_prefix='SEQ',
    stop_values=None
):
    """
    Generic function to process table data for a single sheet
    Returns list of INSERT statements for main table and optional logic table
    Uses global wb, sheetnames, and table_info instead of loading files
    """
    if stop_values is None:
        stop_values = STOP_VALUES

    global wb, sheetnames, table_info
    columns_info = table_info.get(table_name, [])
    logic_columns_info = table_info.get(logic_table_name, []) if logic_table_name else []

    if sheet_idx >= len(sheetnames):
        return []

    ws = wb[sheetnames[sheet_idx]]
    insert_statements = []
    seq_counter = 1

    print(f"  Processing {table_name} data for sheet {sheet_idx}: {sheetnames[sheet_idx]}")
    
    # Pre-calculate column names for better performance
    column_names = [col_info.get('COLUMN_NAME', '') for col_info in columns_info]
    columns_str = ", ".join(column_names)
    
    # Batch data collection
    batch_data = []
    
    # Scan from top to bottom for cell_b_value
    logic_processed = False  # Flag to track if logic has been processed for current main entry
    
    for row_num in range(1, ws.max_row + 1):
        cell_b = ws[f"B{row_num}"]
        if cell_b.value == cell_b_value:
            check_row = row_num + 1
            logic_processed = False
            while check_row <= ws.max_row:
                should_stop = should_stop_row(ws, check_row, stop_values, cell_b_value)
                if should_stop == 'stop':
                    # Create INSERT statements from batch
                    if batch_data:
                        for values in batch_data:
                            values_str = ", ".join(str(v) for v in values)
                            sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
                            insert_statements.append(sql)
                    return insert_statements
                elif should_stop == 'skip':
                    check_row += 1
                    continue
                elif should_stop == 'continue':
                    current_seq = seq_counter
                    row_values = []
                    aoji_values = []  # Collect aoji values from all columns
                    
                    for col_info in columns_info:
                        if column_value_processor:
                            val, aoji = column_value_processor(col_info, ws, check_row, sheet_seq, current_seq)
                        else:
                            val, aoji = set_value_generic(col_info, ws, check_row, sheet_seq, current_seq)
                        row_values.append(val)
                        aoji_values.append(aoji)
                    
                    # Check if there's an AOJI column and set its value based on collected aoji values
                    final_aoji = any(aoji_values)  # True if any column has non-black font
                    for i, col_name in enumerate(column_names):
                        if col_name == 'AOJI':
                            row_values[i] = "'1'" if final_aoji else "'0'"
                            break
                    
                    # Handle MIDASHI special case
                    if len(row_values) > 0:
                        # Find MIDASHI column index
                        midashi_idx = None
                        special_cols_indices = []
                        for i, col_name in enumerate(column_names):
                            if col_name == 'MIDASHI':
                                midashi_idx = i
                            elif col_name in ['IN_GAMEN_ID', 'IN_GAMEN_NAME', 'IN_BUHIN_CD', 'IN_BUHIN_NAME', 'OUT_BUHIN_CD', 'OUT_BUHIN_NAME', 'BIKOU']:
                                special_cols_indices.append(i)
                        
                        if midashi_idx is not None and row_values[midashi_idx] == "'True'":
                            for idx in special_cols_indices:
                                row_values[idx] = 'NULL'
                    logic_processed = False
                    batch_data.append(row_values)
                    seq_counter += 1
                    print(f"    Created {table_name.split('_')[-1]} with Sheet SEQ {sheet_seq} {seq_prefix} {current_seq} at row {check_row}")
                    check_row += 1
                    continue
                elif should_stop == 'create_logic' and logic_table_name and logic_processor and not logic_processed:
                    # Process logic table if provided and logic processor available
                    # Only process logic once per main entry
                    logic_inserts, logic_end_row = logic_processor(
                        ws, check_row, sheet_seq, seq_counter - 1, logic_columns_info
                    )
                    insert_statements.extend(logic_inserts)
                    logic_processed = True  # Mark logic as processed
                    # Skip to end of logic section to avoid reprocessing
                    check_row = logic_end_row if logic_end_row > check_row else check_row + 1
                    continue
                elif should_stop == 'create_logic' and logic_processed:
                    # Logic already processed, just skip this row
                    check_row += 1
                    continue
                else:
                    check_row += 1
    
    # Create INSERT statements from remaining batch data
    if batch_data:
        for values in batch_data:
            values_str = join_sql_values(values)
            sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
    
    return insert_statements


def _get_processor_function(processor_name):
    """Get processor function by name"""
    processor_map = {
        'koumoku_set_value': koumoku_set_value,
        'func_set_value': func_set_value,
        'csv_set_value': csv_set_value,
        'message_set_value': message_set_value,
        'youken_set_value': youken_set_value,
        'hyouji_set_value': hyouji_set_value,
        'tab_set_value': tab_set_value,
        'ichiran_set_value': ichiran_set_value,
        'menu_set_value': menu_set_value,
        'ipo_set_value': ipo_set_value,
        'koumoku_logic': koumoku_logic,
        'func_logic': func_logic,
        'csv_logic': csv_logic,
        're_set_value': re_set_value,
        're_logic': re_logic
    }
    return processor_map.get(processor_name)

def create_row_processor(processor_type):
    """Factory function to create row processors"""
    if processor_type not in ROW_PROCESSOR_CONFIG:
        raise ValueError(f"Unknown processor type: {processor_type}")
    
    config = ROW_PROCESSOR_CONFIG[processor_type]
    
    def row_processor(sheet_idx, sheet_seq, stop_values=None, cell_b_value=None):
        """Generic row processor function"""
        actual_cell_b_value = cell_b_value or config['cell_b_value']
        column_processor = _get_processor_function(config['column_value_processor'])
        logic_processor = None
        
        if 'logic_processor' in config:
            logic_processor = _get_processor_function(config['logic_processor'])
        
        return gen_row_single_sheet(
            sheet_idx=sheet_idx,
            sheet_seq=sheet_seq,
            table_name=config['table_name'],
            logic_table_name=config.get('logic_table_name'),
            cell_b_value=actual_cell_b_value,
            column_value_processor=column_processor,
            logic_processor=logic_processor,
            seq_prefix=config['seq_prefix'],
            stop_values=stop_values
        )
    
    return row_processor

def create_logic_processor(processor_type):
    """Factory function to create logic processors"""
    if processor_type not in LOGIC_PROCESSOR_CONFIG:
        raise ValueError(f"Unknown logic processor type: {processor_type}")
    
    config = LOGIC_PROCESSOR_CONFIG[processor_type]
    
    def logic_processor(ws, start_row, sheet_seq, parent_seq_value, logic_columns_info, cell_b_value=None):
        """Generic logic processor function"""
        actual_cell_b_value = cell_b_value or config['cell_b_value']
        column_processor = _get_processor_function(config['column_value_processor'])
        
        return logic_data_generic(
            ws=ws,
            start_row=start_row,
            sheet_seq=sheet_seq,
            parent_seq_value=parent_seq_value,
            logic_columns_info=logic_columns_info,
            table_name=config['table_name'],
            column_value_processor=column_processor,
            seq_counter_name=config['seq_counter_name'],
            cell_b_value=actual_cell_b_value
        )
    
    return logic_processor

# Create all row processors
koumoku_row = create_row_processor('koumoku')
func_row = create_row_processor('func')
re_row = create_row_processor('re')
csv_row = create_row_processor('csv')
message_row = create_row_processor('message')
tab_row = create_row_processor('tab')
hyouji_row = create_row_processor('hyouji')
ichiran_row = create_row_processor('ichiran')
menu_row = create_row_processor('menu')
ipo_row = create_row_processor('ipo')
youken_row = create_row_processor('youken')

# Create all logic processors
koumoku_logic = create_logic_processor('koumoku_logic')
func_logic = create_logic_processor('func_logic')
re_logic = create_logic_processor('re_logic')
csv_logic = create_logic_processor('csv_logic')


def logic_data_generic(
    ws, 
    start_row, 
    sheet_seq, 
    parent_seq_value, 
    logic_columns_info,
    table_name,
    column_value_processor,
    seq_counter_name,
    stop_values=None,
    cell_b_value=None
):
    """
    Generic function to process logic table data with performance optimizations
    """
    if stop_values is None:
        stop_values = STOP_VALUES
        
    insert_statements = []
    seq_counter = 1
    last_processed_row = start_row
    
    # Pre-calculate column names and table type for better performance
    column_names = [col_info.get('COLUMN_NAME', '') for col_info in logic_columns_info]
    columns_str = ", ".join(column_names)
    logic_type = table_name.split('_')[-1]  # Extract LOGIC type name
    
    # Pre-load cell values for the range
    preload_sheet_cell_values(ws, start_row, min(start_row + 100, ws.max_row), ['B', 'C', 'D', 'E'])
    
    # Batch data collection
    batch_data = []
    
    for check_row in range(start_row, ws.max_row + 1):
        # Use appropriate stopping condition
        should_stop = should_stop_logic_row(ws, check_row, stop_values, cell_b_value)
        if should_stop == 'stop':
            # Process remaining batch
            if batch_data:
                for values in batch_data:
                    values_str = join_sql_values(values)
                    sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
                    insert_statements.append(sql)
            return insert_statements, check_row
        elif should_stop == 'skip':
            continue
        elif should_stop == 'continue':
            # Collect row data
            row_values = []
            aoji_values = []
            for col_info in logic_columns_info:
                val, aoji = column_value_processor(col_info, ws, check_row, sheet_seq, parent_seq_value, seq_counter)
                row_values.append(val)
                aoji_values.append(aoji)
            
            # Check if there's an AOJI column and set its value based on collected aoji values
            final_aoji = any(aoji_values)  # True if any column has non-black font
            for i, col_name in enumerate(column_names):
                if col_name == 'AOJI':
                    row_values[i] = "'1'" if final_aoji else "'0'"
                    break
            
            # Handle YOUKEN_NO special case
            if len(row_values) > 0:
                youken_no_idx = None
                youken_logic_idx = None
                for i, col_name in enumerate(column_names):
                    if col_name == 'YOUKEN_NO':
                        youken_no_idx = i
                    elif col_name == 'YOUKEN_LOGIC':
                        youken_logic_idx = i
                
                if (youken_no_idx is not None and youken_logic_idx is not None and 
                    row_values[youken_no_idx] not in [None, '', "''"]):
                    row_values[youken_logic_idx] = "''"
            
            batch_data.append(row_values)
            print(f"      Created {logic_type} with Sheet SEQ {sheet_seq} Parent SEQ {parent_seq_value} {seq_counter_name} {seq_counter} at row {check_row}")
            seq_counter += 1
            last_processed_row = check_row
    
    # Process remaining batch
    if batch_data:
        for values in batch_data:
            values_str = join_sql_values(values)
            sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str});"
            insert_statements.append(sql)
    
    return insert_statements, last_processed_row


if __name__ == "__main__":
    print("Starting processing all tables in sequence...")
    all_inserts = all_tables_in_sequence('docX.xlsx', 'table_info.txt', 'insert_all.sql')
    print(f"Generated {len(all_inserts)} INSERT statements in total.")


